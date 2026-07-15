#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_lista_xlsx.py — gera a planilha mensal "Listagem de Alunos Ativos" a partir
do data/freq_multi.json (produzido pelo build_freq_multi.py).

Colunas: Nome | Email | Telefone | Categoria | Unidade  (+ Matrícula, Modalidade)
Categoria = comportamento (cluster) IDÊNTICO ao dashboard de Frequência
            (via scripts/cluster_port.py, janela completa = como o dash abre).

Uso: python3 export_lista_xlsx.py [data_dir] [saida.xlsx]
"""
import os, sys, json, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cluster_port import categoria_aluno

MESES_PT = ["", "Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho",
            "Agosto","Setembro","Outubro","Novembro","Dezembro"]
CAT_ORDER = ["Fiel","Regular","Em ascensão","Novo","Retorno","Esporádico",
             "Em declínio","Em evasão","Sem acesso"]

def main():
    data_dir = sys.argv[1] if len(sys.argv) > 1 else "data"
    fj = os.path.join(data_dir, "freq_multi.json")
    d = json.load(open(fj, encoding="utf-8"))
    students = d.get("students", [])
    meses = d.get("meses", [])
    n_meses = len(meses)
    unidades = d.get("unidades", [])
    ulabel = {u.get("key"): u.get("label", u.get("key")) for u in unidades}

    hoje = datetime.date.today()
    mes_ref = f"{MESES_PT[hoje.month]}/{hoje.year}"
    ym = hoje.strftime("%Y-%m")

    linhas = []
    for s in students:
        cat = categoria_aluno(s, n_meses)
        linhas.append({
            "nome": s.get("nome",""), "email": s.get("email",""),
            "tel": s.get("tel",""), "categoria": cat,
            "unidade": ulabel.get(s.get("u"), s.get("u","")),
            "mat": s.get("mat",""), "modalidade": s.get("grupo",""),
        })
    linhas.sort(key=lambda x: (x["unidade"], x["nome"]))

    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        data_dir, f"Listagem Alunos Ativos ({ym}).xlsx")
    build_xlsx(linhas, mes_ref, ym, out)
    sem_email = sum(1 for l in linhas if not l["email"])
    sem_tel = sum(1 for l in linhas if not l["tel"])
    print(f"[ok] {out} | {len(linhas)} ativos | sem e-mail: {sem_email} | sem telefone: {sem_tel}")
    return out

def build_xlsx(linhas, mes_ref, ym, path):
    AZUL, CINZA, BRANCO, FONT = "1F3864", "F2F2F2", "FFFFFF", "Arial"
    f_title=Font(name=FONT,size=16,bold=True,color=BRANCO); f_sub=Font(name=FONT,size=10,color=BRANCO)
    f_head=Font(name=FONT,size=11,bold=True,color=BRANCO); f_cell=Font(name=FONT,size=10)
    f_bold=Font(name=FONT,size=10,bold=True); f_note=Font(name=FONT,size=9,italic=True,color="7F7F7F")
    fill_t=PatternFill("solid",fgColor=AZUL); fill_h=PatternFill("solid",fgColor=AZUL)
    fill_leg=PatternFill("solid",fgColor="D9E1F2")
    thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    center=Alignment(horizontal="center",vertical="center",wrap_text=True)
    left=Alignment(horizontal="left",vertical="center",wrap_text=True)
    wb=openpyxl.Workbook()

    ws=wb.active; ws.title="Instruções"; ws.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=26; ws.column_dimensions["C"].width=84
    ws.merge_cells("B2:C2"); ws["B2"]="Nad'Arte — Listagem de Alunos Ativos"
    ws["B2"].font=f_title; ws["B2"].fill=fill_t; ws["B2"].alignment=left
    ws.merge_cells("B3:C3"); ws["B3"]=f"Mês de referência: {mes_ref} · {len(linhas)} alunos ativos"
    ws["B3"].font=f_sub; ws["B3"].fill=fill_t; ws["B3"].alignment=left; ws.row_dimensions[2].height=30
    info=[("Fonte","Pacto (ERP) via API — nome, e-mail, telefone, unidade, matrícula, modalidade."),
          ("Categoria","Comportamento (cluster) idêntico ao dashboard de Frequência."),
          ("Rótulos","Fiel · Regular · Em ascensão · Novo · Retorno · Esporádico · Em declínio · Em evasão · Sem acesso."),
          ("Ativo do mês","Base ativa do Pacto no mês de referência (mesmo critério do dash)."),
          ("Gerado em",hoje_str())]
    r=5
    for t,tx in info:
        ws.cell(r,2,t).font=f_bold; ws.cell(r,2).fill=fill_leg; ws.cell(r,2).alignment=left; ws.cell(r,2).border=border
        ws.cell(r,3,tx).font=f_cell; ws.cell(r,3).alignment=left; ws.cell(r,3).border=border; ws.row_dimensions[r].height=22; r+=1

    ws2=wb.create_sheet("Alunos"); ws2.sheet_view.showGridLines=False
    COLS=["Nome","Email","Telefone","Categoria","Unidade","Matrícula","Modalidade"]
    for i,w in enumerate([28,32,20,15,15,14,18],1): ws2.column_dimensions[get_column_letter(i)].width=w
    ws2.merge_cells("A1:G1"); ws2["A1"]=f"Alunos ativos — {mes_ref}"; ws2["A1"].font=f_title; ws2["A1"].fill=fill_t; ws2["A1"].alignment=left; ws2.row_dimensions[1].height=26
    for i,c in enumerate(COLS,1):
        cell=ws2.cell(2,i,c); cell.font=f_head; cell.fill=fill_h; cell.alignment=center; cell.border=border
    r=3
    for a in linhas:
        for i,v in enumerate([a["nome"],a["email"],a["tel"],a["categoria"],a["unidade"],a["mat"],a["modalidade"]],1):
            cell=ws2.cell(r,i,v); cell.font=f_cell; cell.border=border; cell.alignment=(left if i<=2 else center)
        r+=1
    ultima=max(r-1,2); ws2.freeze_panes="A3"; ws2.auto_filter.ref=f"A2:G{ultima}"

    ws3=wb.create_sheet("Resumo"); ws3.sheet_view.showGridLines=False
    unidades=sorted({a["unidade"] for a in linhas})
    cats=[c for c in CAT_ORDER if any(a["categoria"]==c for a in linhas)] or CAT_ORDER
    ws3.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cats)+2)
    ws3["A1"]=f"Resumo — alunos por unidade e categoria ({mes_ref})"; ws3["A1"].font=f_title; ws3["A1"].fill=fill_t; ws3["A1"].alignment=left; ws3.row_dimensions[1].height=26
    ws3.column_dimensions["A"].width=16
    for i in range(2,3+len(cats)): ws3.column_dimensions[get_column_letter(i)].width=13
    hr=3
    ws3.cell(hr,1,"Unidade").font=f_head; ws3.cell(hr,1).fill=fill_h; ws3.cell(hr,1).alignment=center; ws3.cell(hr,1).border=border
    for j,c in enumerate(cats,2):
        cell=ws3.cell(hr,j,c); cell.font=f_head; cell.fill=fill_h; cell.alignment=center; cell.border=border
    totcol=2+len(cats); cell=ws3.cell(hr,totcol,"Total"); cell.font=f_head; cell.fill=fill_h; cell.alignment=center; cell.border=border
    counts={u:{c:0 for c in cats} for u in unidades}
    for a in linhas:
        if a["categoria"] in counts[a["unidade"]]: counts[a["unidade"]][a["categoria"]]+=1
    r=hr+1
    for u in unidades:
        ws3.cell(r,1,u).font=f_bold; ws3.cell(r,1).border=border; ws3.cell(r,1).alignment=left
        for j,c in enumerate(cats,2):
            cell=ws3.cell(r,j,counts[u][c]); cell.font=f_cell; cell.border=border; cell.alignment=center
        first,last=get_column_letter(2),get_column_letter(1+len(cats))
        cell=ws3.cell(r,totcol,f"=SUM({first}{r}:{last}{r})"); cell.font=f_bold; cell.border=border; cell.alignment=center
        r+=1
    ws3.cell(r,1,"Total").font=f_head; ws3.cell(r,1).fill=fill_h; ws3.cell(r,1).alignment=center; ws3.cell(r,1).border=border
    for j in range(2,totcol+1):
        col=get_column_letter(j); cell=ws3.cell(r,j,f"=SUM({col}{hr+1}:{col}{r-1})"); cell.font=f_head; cell.fill=fill_h; cell.border=border; cell.alignment=center
    wb.save(path)

def hoje_str():
    return datetime.date.today().strftime("%d/%m/%Y")

hoje = datetime.date.today()
if __name__ == "__main__":
    main()
