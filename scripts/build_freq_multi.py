#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Motor de reconstrucao do freq_multi.json — Dashboard Frequencia & Retencao Nad'Arte.
Implementa secoes 4-8 do runbook. Le os .bin (xlsx decodificados) listados em data/manifest.json
e produz data/freq_multi.json (JSON valido, minificado, ensure_ascii=False).

Janela de meses, mes-base, faixa-etaria de referencia e junMax sao DINAMICOS
(derivados dos arquivos), entao o pipeline segue valendo quando entrar Jul, Ago, etc.

Uso: python3 build_freq_multi.py <data_dir>
  <data_dir> contem os arquivos <fileId>.bin e manifest.json (id -> titulo).
"""
import sys, os, io, re, json, unicodedata, datetime, calendar
from collections import defaultdict, Counter
import openpyxl

DATA_DIR = sys.argv[1] if len(sys.argv) > 1 else "data"

# ---- constantes do runbook ----
SHEET_TO_UNIT = {"Lago Norte":"LagoNorte","905 Sul":"905Sul","604 Norte":"604Norte",
                 "716 Norte":"716Norte","Lago Sul":"LagoSul"}
UDPS = {"716Norte":7,"905Sul":7,"604Norte":6,"LagoNorte":6,"LagoSul":6}
UNIDADES = [{"key":"REDE","label":"Rede (todas)"},
            {"key":"716Norte","label":"716 Norte"},{"key":"905Sul","label":"905 Sul"},
            {"key":"604Norte","label":"604 Norte"},{"key":"LagoNorte","label":"Lago Norte"},
            {"key":"LagoSul","label":"Lago Sul"}]
TICKETS = {"716Norte":275.00,"905Sul":268.00,"604Norte":273.80,"LagoNorte":299.60,"LagoSul":266.60}
TICKET_NATAL = 245.20
UNIT_KEYS = ["716Norte","905Sul","604Norte","LagoNorte","LagoSul"]
# Natal (RN): unidade NOVA (catraca desde 01/07/2026). DESLIGADA ate a Variable PACTO_ENABLE_NATAL=1
# (fica fora da Rede/seletor/churn ate ligar). So aparece de jul/26 em diante (trava no pacto_fetch).
if os.environ.get("PACTO_ENABLE_NATAL") == "1":
    SHEET_TO_UNIT["Natal (RN)"] = "Natal"
    UDPS["Natal"] = 6                        # seg-sab (domingo fechado)
    UNIDADES.append({"key":"Natal","label":"Natal (RN)"})
    TICKETS["Natal"] = 276.60                # PLACEHOLDER (media da rede) — CONFIRMAR ticket real antes do go-live
    UNIT_KEYS.append("Natal")
ABBR = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}

# --- Classificador de modalidade (texto do plano -> 7 baldes) ---
# Fonte: guia "Pacto API - Modalidades e Categorias". A modalidade vem da descricao
# do contrato (GET /v1/contrato/matricula/{mat}), nao do cadastro (que vem vazio).
AGUA_TK = ("NATAC","NATA","HIDRO","BEBE","AQUA")
LUTA_TK = ("KARATE","MUAY","JIU","JUDO","HAPKIDO","CAPOEIRA","BOXE","TAEKWON","KUNG","LUTA")
FIT_TK  = ("TRANSITO LIVRE"," TL ","LIVRE ACESSO","FITNESS","MUSCULA","DANCA","PILATES",
           "AULA COLETIVA","FUNCIONAL","SPINNING","CROSS","ZUMBA","RITMO","GINASTICA",
           "ALONGA","YOGA","TREINA")

# Rotulos canonicos das categorias (identicos aos usados no dashboard)
CAT_FIT="Fitness"; CAT_AGUA="Água"; CAT_LUTA="Luta"
CAT_AF="Ambos (Água + Fitness)"; CAT_AL="Ambos (Água + Luta)"
CAT_FL="Ambos (Fitness + Luta)"; CAT_AFL="Ambos (Água + Fitness + Luta)"
CAT_OUT="Outros"

# LAGO NORTE: a catraca fica so no ambiente fitness. So passam pela catraca as categorias
# que incluem Fitness; alunos exclusivos de Agua/Luta ficam de fora da analise de frequencia.
LAGO_UNIT = "LagoNorte"
LAGO_FIT_CATS = {CAT_FIT, CAT_AF, CAT_FL, CAT_AFL}
LAGO_EXCLUDED = set()   # (unit,key) Lago Norte sem catraca (Agua/Luta) — usado no aviso/contagem do Lago
FREQ_BLIND = set()      # (unit,key) CEGOS DE CATRACA no geral: entram no CHURN, saem da FREQUENCIA
# Natal (RN): unidade com histórico de contrato (churn ok), mas catraca so a partir de 01/07/2026.
# Ate a catraca ter massa, a unidade INTEIRA fica cega de catraca (churn sim, frequencia nao).
# Liga a frequencia com PACTO_NATAL_FREQ=1 (previsto ~ago/26, quando julho fechar).
NATAL_UNIT = "Natal"
NATAL_FREQ = os.environ.get("PACTO_NATAL_FREQ", "0") == "1"

def deaccent(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
def up(s):
    return deaccent(str(s or "")).upper().strip()
def norm_mat(v):
    d = re.sub(r"\D","", str(v if v is not None else "")); d = d.lstrip("0")
    return d if d else ("0" if re.search(r"\d", str(v or "")) else "")
def norm_doc(v):
    return re.sub(r"\D","", str(v if v is not None else ""))
def norm_nome(v):
    return up(v)
NAME2CPF = {}  # (unit, nome_normalizado) -> cpf  (ponte: resolve linhas/meses sem CPF)
def skey(unit, doc_val, nome_val):
    # chave canonica: CPF quando conhecido (na linha OU via ponte por nome); senao Nome. Sempre por unidade.
    cpf = norm_doc(doc_val); n = norm_nome(nome_val)
    if len(cpf) < 11 and n:
        cpf = NAME2CPF.get((unit, n), "")
    if len(cpf) >= 11:
        return f"{unit}|C|{cpf}"
    return f"{unit}|N|{n}" if n else ""

def _tok_cat(tok):
    # prioridade agua > fit > luta; padding com espacos evita casar "TL" dentro de palavra
    t = " " + up(tok) + " "
    if any(k in t for k in AGUA_TK): return "agua"
    if any(k in t for k in FIT_TK):  return "fit"
    if any(k in t for k in LUTA_TK): return "lutas"
    return "outros"
def classify_grupo(mod):
    # texto da(s) descricao(oes) de contrato -> 7 baldes (+ Outros). Tokeniza por ; , + /
    u = up(mod)
    if "TODAS AS MODALIDADES" in u or "TODAS MODALIDADES" in u:
        return CAT_AFL
    toks = [t for t in re.split(r"[;,+/]", str(mod or "")) if t.strip()]
    b = set(_tok_cat(t) for t in toks); b.discard("outros")
    A = "agua" in b; F = "fit" in b; L = "lutas" in b
    if A and F and L: return CAT_AFL
    if A and F: return CAT_AF
    if A and L: return CAT_AL
    if F and L: return CAT_FL
    if A: return CAT_AGUA
    if F: return CAT_FIT
    if L: return CAT_LUTA
    return CAT_OUT

def parse_birth(v):
    if v is None or v == "": return (0,0,0)
    if isinstance(v,(datetime.datetime,datetime.date)): return (v.month, v.day, v.year)
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if m:
        d,mo,y = int(m.group(1)),int(m.group(2)),int(m.group(3))
        if y < 100: y += 2000 if y < 50 else 1900
        return (mo,d,y)
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m: return (int(m.group(2)), int(m.group(3)), int(m.group(1)))
    return (0,0,0)

def band_of(ref, by,bm,bd):
    if not by: return "N/D"
    age = ref.year - by - ((ref.month, ref.day) < (bm, bd))
    if age <= 11: return "0–11"
    if age <= 17: return "12–17"
    if age <= 29: return "18–29"
    if age <= 44: return "30–44"
    if age <= 59: return "45–59"
    return "60+"
def sexo_of(v):
    s = up(v)
    if s.startswith("M"): return "M"
    if s.startswith("F"): return "F"
    return "N/D"
def load_wb(path):
    return openpyxl.load_workbook(io.BytesIO(open(path,"rb").read()), read_only=True, data_only=True)
def detect_header(rows, tokens, scan=8):
    for i in range(min(scan, len(rows))):
        cells = rows[i]; names = [up(c) for c in cells]
        if sum(1 for tk in tokens if any(tk in n for n in names)) >= 2:
            colmap = {up(c):ci for ci,c in enumerate(cells) if c is not None and str(c).strip()!=""}
            return i, colmap
    return None, None
def find_col(colmap, *cands):
    # 1) match EXATO (evita que "MATRICULA" capture "DATA MATRICULA", etc.)
    for cand in cands:
        cu = up(cand)
        if cu in colmap: return colmap[cu]
    # 2) match por substring (tolerante a variacoes de cabecalho)
    for cand in cands:
        cu = up(cand)
        for name,ci in colmap.items():
            if cu in name: return ci
    return None
# situacaoContrato (motivo de saida) -> rotulo legivel. Vem cru da Pacto (DESISTENTE, CANCELADO...).
MOTIVO_ROT = {
    "DESISTENTE":"Desistência","CANCELADO":"Cancelamento","INATIVO_VENCIDO":"Contrato vencido",
    "TRANCADO_VENCIDO":"Trancado/vencido","A_VENCER":"A vencer","ATESTADO":"Atestado médico",
    "NORMAL":"Ativo/normal","TRANCADO":"Trancado","":"Sem informação",
}
def motivo_norm(v):
    s = str(v or "").strip().upper()
    if not s or s in ("?","NONE","NULL"): return ""
    return MOTIVO_ROT.get(s, s.replace("_"," ").capitalize())
def parse_dt(v):
    if isinstance(v,(datetime.datetime,datetime.date)):
        return datetime.date(v.year,v.month,v.day)
    s=str(v or "").strip()
    m=re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        try: return datetime.date(int(m.group(1)),int(m.group(2)),int(m.group(3)))
        except ValueError: pass
    m=re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        d,mo,y = int(m.group(1)),int(m.group(2)),int(m.group(3))
        if mo>12 and d<=12: d,mo = mo,d      # tolera MM/DD/AAAA invertido
        try: return datetime.date(y,mo,d)
        except ValueError: pass
    return None

# ---- carregar manifest e classificar arquivos ----
manifest = json.load(open(os.path.join(DATA_DIR,"manifest.json")))
alunos_files = {}   # (year,month) -> path
catraca_files = {}  # unit_key -> path
for fid,title in manifest.items():
    path = os.path.join(DATA_DIR, fid+".bin")
    m = re.match(r"\s*(\d{1,2})\.\s*Alunos Ativos.*\((\w+)\.(\d{4})\)", title)
    if m:
        alunos_files[(int(m.group(3)), int(m.group(1)))] = path; continue
    mc = re.search(r"Acessos Catr?ata Unidade (.+?) \(Nad", title)
    if mc:
        uname = mc.group(1).strip()
        catraca_files[SHEET_TO_UNIT.get(uname, uname.replace(" ",""))] = path
assert alunos_files and catraca_files, "faltam arquivos de alunos ou catraca"

# timeline de meses = uniao (alunos + catraca), ordenada
month_keys = set(alunos_files.keys())
SHEET_MONTH = re.compile(r"^\s*(\d{1,2})\.\s*\w+\.?(\d{4})?")
for unit,path in catraca_files.items():
    for sn in load_wb(path).sheetnames:
        mm=SHEET_MONTH.match(sn)
        if mm:
            yr=int(mm.group(2)) if mm.group(2) else max(y for (y,_) in alunos_files)
            month_keys.add((yr,int(mm.group(1))))
ORDERED = sorted(month_keys)
POS = {mk:i for i,mk in enumerate(ORDERED)}
NMONTHS = len(ORDERED)
MESES = [f"{ABBR[mn]}.{str(yr)[2:]}" for (yr,mn) in ORDERED]
base_key = max(alunos_files.keys())
base_pos = POS[base_key]
by_, bm_ = base_key
BAND_REF = datetime.date(by_, bm_, calendar.monthrange(by_, bm_)[1])  # ultimo dia do mes-base
print(f"[info] meses={MESES} base={MESES[base_pos]} bandref={BAND_REF}", file=sys.stderr)

# ---- pre-pass: ponte nome->CPF (alguns meses/unidades vem sem a coluna CPF) ----
for mk, path in alunos_files.items():
    wb = load_wb(path)
    for sn in wb.sheetnames:
        unit = SHEET_TO_UNIT.get(sn.strip())
        if not unit: continue
        rows = [r for r in wb[sn].iter_rows(values_only=True)]
        hidx, colmap = detect_header(rows, ["MATRICULA","DOCUMENTO","NASCIMENTO"])
        if hidx is None: continue
        cD=find_col(colmap,"DOCUMENTO","CPF"); cN=find_col(colmap,"NOME")
        if cD is None or cN is None: continue
        for r in rows[hidx+1:]:
            if r is None: continue
            cpf = norm_doc(r[cD]) if cD<len(r) else ""
            nome = norm_nome(r[cN]) if (cN<len(r) and r[cN] is not None) else ""
            if len(cpf)>=11 and nome:
                NAME2CPF.setdefault((unit,nome), cpf)
print(f"[info] ponte nome->CPF: {len(NAME2CPF)} nomes mapeados", file=sys.stderr)

def plano_of(ini_iso, fim_iso):
    """Prazo do plano pelo intervalo inicio->fim do contrato -> (rotulo, meses padrao).
    Sem uma das datas, retorna ('?',0). Arredonda pro plano padrao mais proximo."""
    di=parse_dt(ini_iso); df=parse_dt(fim_iso)
    if not di or not df: return ("?", 0)
    m=(df.year-di.year)*12 + (df.month-di.month)   # diff de meses (limite exato: Jul->Ago = 1)
    if m<=0: return ("?", 0)
    cand=[(1,"Mensal"),(2,"Bimestral"),(3,"Trimestral"),(6,"Semestral"),(12,"Anual")]
    meses,rot=min(((mm,rr) for mm,rr in cand), key=lambda c:abs(c[0]-m))
    return (rot, meses)

# ---- parse alunos: active[(pos,unit)] = set(mat); attrs[pos][(unit,mat)] ----
active = defaultdict(set); attrs = defaultdict(dict)
for mk, path in alunos_files.items():
    pos = POS[mk]; wb = load_wb(path)
    for sn in wb.sheetnames:
        unit = SHEET_TO_UNIT.get(sn.strip())
        if not unit: continue
        rows = [r for r in wb[sn].iter_rows(values_only=True)]
        hidx, colmap = detect_header(rows, ["MATRICULA","DOCUMENTO","NASCIMENTO"])
        if hidx is None: continue
        c_mat=find_col(colmap,"MATRICULA"); c_nome=find_col(colmap,"NOME")
        c_doc=find_col(colmap,"DOCUMENTO","CPF")
        c_nasc=find_col(colmap,"NASCIMENTO"); c_sexo=find_col(colmap,"SEXO"); c_mod=find_col(colmap,"MODALIDADE")
        c_dm=find_col(colmap,"DATA MATRICULA","DATA DE MATRICULA","DT MATRICULA","DATA MATR")
        c_foto=find_col(colmap,"FOTO"); c_prof=find_col(colmap,"PROF NOME","PROFESSOR"); c_prole=find_col(colmap,"PROF TIPO","PROF FOTO")
        c_venc=find_col(colmap,"VENCIMENTO","FIM CONTRATO","DATA FIM","VENC","FIM")
        c_ini=find_col(colmap,"INICIO CONTRATO","INICIO DO CONTRATO")
        c_email=find_col(colmap,"EMAIL","E-MAIL"); c_tel=find_col(colmap,"TELEFONE","CELULAR","FONE")
        c_motivo=find_col(colmap,"MOTIVO SAIDA","SITUACAO CONTRATO","SITUACAOCONTRATO")
        for r in rows[hidx+1:]:
            if r is None: continue
            nome_v = str(r[c_nome]).strip() if (c_nome is not None and c_nome<len(r) and r[c_nome] is not None) else ""
            doc_v = r[c_doc] if (c_doc is not None and c_doc<len(r)) else None
            mat = norm_mat(r[c_mat]) if (c_mat is not None and c_mat<len(r)) else ""
            key = mat if mat else skey(unit, doc_v, nome_v)   # chave = (unidade+matricula); fallback CPF/nome
            if not key: continue
            mod = r[c_mod] if (c_mod is not None and c_mod<len(r)) else ""
            grupo = classify_grupo(mod)
            # LAGO NORTE: catraca so no acesso fitness. Aluno sem Fitness (Agua/Luta puros) nao
            # passa pela catraca. CEGOS DE CATRACA: entram no churn/base/retencao (medidos por
            # CONTRATO, sem depender de acesso), mas ficam FORA dos sinais de FREQUENCIA
            # (senao virariam falso "sumido"). So marca aqui; segue e entra em active/attrs.
            if unit == LAGO_UNIT and grupo not in LAGO_FIT_CATS:
                LAGO_EXCLUDED.add((unit, key)); FREQ_BLIND.add((unit, key))   # Lago sem catraca (Agua/Luta)
            elif unit == NATAL_UNIT and not NATAL_FREQ:
                FREQ_BLIND.add((unit, key))   # Natal: catraca nova (jul/26) sem massa -> churn sim, frequencia nao
            active[(pos,unit)].add(key)
            nasc = r[c_nasc] if (c_nasc is not None and c_nasc<len(r)) else None
            bm,bd,by = parse_birth(nasc)
            _cell=lambda ci: (str(r[ci]).strip() if (ci is not None and ci<len(r) and r[ci] is not None) else "")
            _venc_iso = (parse_dt(r[c_venc]).isoformat() if (c_venc is not None and c_venc<len(r) and parse_dt(r[c_venc])) else "")
            _ini_iso  = (parse_dt(r[c_ini]).isoformat()  if (c_ini  is not None and c_ini<len(r)  and parse_dt(r[c_ini]))  else "")
            _plano_rot, _plano_m = plano_of(_ini_iso, _venc_iso)   # mensal/bi/tri/sem/anual pelo prazo real
            attrs[pos][(unit,key)] = {
                "nome": nome_v, "mat": mat,
                "grupo": grupo, "mod": str(mod or "").strip(),
                "sexo": sexo_of(r[c_sexo]) if (c_sexo is not None and c_sexo<len(r)) else "N/D",
                "band": band_of(BAND_REF,by,bm,bd), "bm":bm,"bd":bd,"by":by,
                "dm": (parse_dt(r[c_dm]).isoformat() if (c_dm is not None and c_dm<len(r) and parse_dt(r[c_dm])) else ""),
                "venc": _venc_iso, "ini": _ini_iso, "plano": _plano_rot, "termoMeses": _plano_m,
                "foto": _cell(c_foto), "prof": _cell(c_prof), "profRole": _cell(c_prole),
                "email": _cell(c_email), "tel": _cell(c_tel),
                "motivo": motivo_norm(_cell(c_motivo)),
            }

# ---- parse catraca: acc[(unit,pos)][mat]; ancora; junMax (ult. data do mes-base) ----
acc = defaultdict(lambda: defaultdict(int)); anchor = {}; max_base_date=None
first_acc = {}  # key -> 1a data de acesso (catraca) de todos os tempos na janela
# ---- Onda 1a: recencia (ultima visita) + acessos por SEMANA (ultimas WEEKS_KEEP semanas) ----
# Sem custo extra de API: as datas ja estao na catraca; so agregamos por semana e guardamos a ultima.
last_acc = {}   # key -> ultima data de acesso (recencia)
def _monday(d): return d - datetime.timedelta(days=d.weekday())
REF_MON = _monday(datetime.date.today())                       # segunda-feira da semana corrente (build)
WEEKS_KEEP = int(os.environ.get("PACTO_WEEKS_KEEP", "10"))     # janela semanal retida por aluno
week_acc = defaultdict(lambda: [0]*WEEKS_KEEP)                 # key -> [contagens] (wk[-1] = semana atual)
for unit, path in catraca_files.items():
    wb = load_wb(path); total=0; people=set()
    for sn in wb.sheetnames:
        mm=SHEET_MONTH.match(sn)
        if not mm: continue
        yr=int(mm.group(2)) if mm.group(2) else by_
        mk=(yr,int(mm.group(1)))
        if mk not in POS: continue
        pos=POS[mk]
        rows = [r for r in wb[sn].iter_rows(values_only=True)]
        hidx, colmap = detect_header(rows, ["MAT. CLIENTE","CPF","DATA ENTRADA"])
        if hidx is None: continue
        c_mat=find_col(colmap,"MAT. CLIENTE","MAT CLIENTE"); c_cpf=find_col(colmap,"CPF"); c_nome=find_col(colmap,"NOME"); c_dt=find_col(colmap,"DATA ENTRADA")
        if c_mat is None and c_cpf is None and c_nome is None: continue
        for r in rows[hidx+1:]:
            if r is None: continue
            mat_v = norm_mat(r[c_mat]) if (c_mat is not None and c_mat<len(r)) else ""
            cpf_v = r[c_cpf] if (c_cpf is not None and c_cpf<len(r)) else None
            nome_v = str(r[c_nome]).strip() if (c_nome is not None and c_nome<len(r) and r[c_nome] is not None) else ""
            key = mat_v if mat_v else skey(unit, cpf_v, nome_v)   # chave = matricula; fallback CPF/nome
            if not key: continue
            acc[(unit,pos)][key]+=1; total+=1; people.add(key)
            if c_dt is not None and c_dt<len(r):
                d=parse_dt(r[c_dt])
                if d:
                    if pos==base_pos and (max_base_date is None or d>max_base_date): max_base_date=d
                    if key not in first_acc or d<first_acc[key]: first_acc[key]=d
                    if key not in last_acc or d>last_acc[key]: last_acc[key]=d
                    _wi = (_monday(d) - REF_MON).days // 7       # 0=semana atual, -1=passada...
                    _idx = WEEKS_KEEP-1 + _wi
                    if 0 <= _idx < WEEKS_KEEP: week_acc[key][_idx]+=1
    anchor[unit]=(total,len(people))
print("[info] catraca anchor:", {u:anchor[u] for u in UNIT_KEYS}, file=sys.stderr)
JUN_MAX = max_base_date.strftime("%d/%m/%Y") if max_base_date else ""
print(f"[info] junMax(ult. data mes-base) = {JUN_MAX}", file=sys.stderr)

# ---- ledger de identidade (perpetuo) + first-seen por aluno: regra de 'novo' SEMPRE pelo historico completo, nunca pela janela filtrada ----
def ym_of(pos):
    yr,mn = ORDERED[pos]; return f"{yr:04d}-{mn:02d}"
YM2POS = {ym_of(i):i for i in range(NMONTHS)}
first_seen_ym = {}
for (pos,unit), keys in active.items():
    slabel = ym_of(pos)
    for k in keys:
        if k not in first_seen_ym or slabel < first_seen_ym[k]:
            first_seen_ym[k] = slabel
# ledger persistente NA RAIZ do repo (commitado pelo workflow) para perpetuar a 1a aparicao entre
# runs -> Novo vs Retorno de verdade. Chaves HASHEADAS (sem PII), seguro em repo publico.
import hashlib as _hl0
def _kh(k): return _hl0.sha256(str(k).encode("utf-8")).hexdigest()[:16]   # hash estavel da chave
LEDGER_PATH = os.environ.get("LEDGER_PATH", "ledger.json")   # raiz (nao gitignorado)
ledger = {}
try: ledger = json.load(open(LEDGER_PATH))
except Exception: ledger = {}
for k, slabel in first_seen_ym.items():
    hk = _kh(k)
    if hk not in ledger or slabel < ledger[hk]:
        ledger[hk] = slabel
try:
    with open(LEDGER_PATH,"w") as f: json.dump(ledger,f,ensure_ascii=False,separators=(",",":"))
except Exception as e:
    print(f"[WARN] nao salvou ledger: {e}", file=sys.stderr)
def fs_index(key):
    ym = ledger.get(_kh(key))
    if not ym: return base_pos
    return YM2POS.get(ym, -1)  # -1 = 1a aparicao ANTES da janela (base arquivada): nunca 'novo'
print(f"[info] ledger: {len(ledger)} pessoas (1a aparicao / perpetuo, hasheado)", file=sys.stderr)

# ==== Onda 2: SCORE DE CHURN no build (FONTE ÚNICA p/ dashboard + Agenda Tática) ====
# Espelha exatamente o scoreOf(s) do template. Constantes calibráveis abaixo (mudar num lugar só).
_TODAY = datetime.date.today()
BASE_PARTIAL = (tuple(ORDERED[base_pos]) == (_TODAY.year, _TODAY.month))
VEZ_META = {CAT_FIT:3, CAT_AGUA:1.5, CAT_LUTA:1.5, CAT_AF:3, CAT_AL:2, CAT_FL:3, CAT_AFL:3, CAT_OUT:1.5}
def _op_cutoff():
    return max_base_date.day if (BASE_PARTIAL and max_base_date) else calendar.monthrange(*ORDERED[base_pos])[1]
def _days_between(a, b):
    if not a or not b: return None
    return (b - a).days
def _meses_desde(dm_iso, today):
    d = parse_dt(dm_iso)
    if not d: return None
    return (today.year - d.year)*12 + (today.month - d.month) - (1 if today.day < d.day else 0)
def score_of(ac, active, last_date, wk, venc_iso, dm_iso, grupo, dps, is_novo, today=None):
    today = today or _TODAY
    reasons=[]; pts=0
    ativo = (active[-1]==1) if active else True
    # 1) aderência do mês corrente (máx 25)
    dias=_op_cutoff(); semanas=max(0.3, dias/dps); expected=VEZ_META.get(grupo,3)*semanas
    acm = ac[base_pos] if base_pos < len(ac) else 0
    ad = min(1.0, acm/expected) if expected>0 else 0.0
    pAd = 0 if ad>=0.9 else 5 if ad>=0.7 else 12 if ad>=0.5 else 18 if ad>=0.3 else 25
    if pAd: reasons.append((pAd, f"Aderência do mês {round(ad*100)}%"))
    pts += pAd
    # 2) recência (máx 25)
    rec = _days_between(last_date, today)
    pRec = 0 if rec is None else 0 if rec<=7 else 12 if rec<=14 else 18 if rec<=21 else 22 if rec<=30 else 25
    if pRec: reasons.append((pRec, f"Recência: {rec} dias sem visita"))
    pts += pRec
    # 3) velocidade de queda (máx 15)
    K=len(wk); vel=None; pVel=0
    if K>=6:
        recent=(wk[K-1]+wk[K-2])/2; parr=wk[max(0,K-8):K-2]; prior=(sum(parr)/len(parr)) if parr else 0
        if prior>0:
            vel=(prior-recent)/prior; pVel = 15 if vel>=0.4 else 10 if vel>=0.2 else 5 if vel>0.05 else 0
    if pVel: reasons.append((pVel, f"Velocidade de queda {round(vel*100)}% (2 últ. sem. vs anteriores)"))
    pts += pVel
    # 4) vencimento (máx 15)
    dv = _days_between(today, parse_dt(venc_iso)); pVenc=0
    if ativo and dv is not None and dv>=0:
        pVenc = 15 if dv<=7 else 12 if dv<=15 else 8 if dv<=30 else 4 if dv<=60 else 0
    if pVenc: reasons.append((pVenc, f"Contrato vence em {dv} dias"))
    pts += pVenc
    # 5) constância (máx 10)
    pCon=0; con=None
    if K>=4:
        wwin=wk[max(0,K-8):]; pres=sum(1 for v in wwin if v>0); con=pres/len(wwin) if wwin else 1
        pCon = 0 if con>=0.8 else 3 if con>=0.6 else 6 if con>=0.4 else 10
    if pCon: reasons.append((pCon, f"Constância {round(con*100)}% das últimas semanas"))
    pts += pCon
    # 6) coorte / tempo de casa (máx 10)
    meses=_meses_desde(dm_iso, today); pCo=0
    if meses is not None:
        pCo = 0 if meses<1 else 6 if meses<=3 else 10 if meses<=6 else 4 if meses<=12 else 2
    if pCo: reasons.append((pCo, f"Tempo de casa: {meses} meses" + (" (janela de evasão)" if (meses>=3 and meses<=6) else "")))
    pts += pCo
    weeks_data = sum(1 for v in wk if v>0)
    conf = "Alta" if (meses is not None and meses>2 and weeks_data>=4) else ("Média" if (meses is not None and meses>=1) else "Baixa")
    if is_novo: conf="Baixa"
    justif = "Contrato inativo (perda realizada, não risco)" if not ativo else ("Matriculado no mês — novo demais p/ classificar" if is_novo else None)
    score = max(0, min(100, round(pts)))
    reasons.sort(key=lambda r: -r[0])
    band = "Justificado" if justif else ("Risco elevado" if score>=70 else "Queda relevante" if score>=50 else "Atenção inicial" if score>=30 else "Dentro do padrão")
    if justif: prio="—"
    elif dv is not None and 0<=dv<=15 and score>=50: prio="P0"
    elif score>=70 and (rec is None or rec<=45): prio="P0"
    elif score>=70: prio="P1"
    elif score>=50: prio="P1"
    elif score>=30: prio="P2"
    else: prio="P3"
    return {"score":score,"band":band,"prio":prio,"conf":conf,"justif":justif,
            "reasons":[{"p":p,"t":t} for p,t in reasons[:5]], "rec":rec, "dv":dv}

# ---- montar students (apenas ativos do mes-base) ----
students=[]
for (pos,unit),keys in list(active.items()):   # snapshot: nao mutar 'active' durante a iteracao
    if pos!=base_pos: continue
    for key in keys:
        if (unit,key) in FREQ_BLIND: continue   # cego de catraca: fora da FREQUENCIA (sem acesso p/ medir)
        a = attrs[base_pos].get((unit,key),{})
        ac = [acc.get((unit,mm),{}).get(key,0) for mm in range(NMONTHS)]
        act= [1 if key in active.get((mm,unit),()) else 0 for mm in range(NMONTHS)]  # .get: nao cria chave em defaultdict
        _dmd = parse_dt(a.get("dm",""))
        _is_novo = bool(_dmd and (_dmd.year,_dmd.month)==tuple(ORDERED[base_pos]))
        _sc = score_of(ac, act, last_acc.get(key), week_acc.get(key,[0]*WEEKS_KEEP), a.get("venc",""), a.get("dm",""), a.get("grupo","Outros"), UDPS[unit], _is_novo)
        students.append({
            "u":unit,"mat":a.get("mat",""),"nome":a.get("nome",""),
            "grupo":a.get("grupo","Outros"),"mod":a.get("mod","")[:40],
            "sexo":a.get("sexo","N/D"),"band":a.get("band","N/D"),"dps":UDPS[unit],
            "bm":a.get("bm"),"bd":a.get("bd"),"by":a.get("by"),"ac":ac,"active":act,"fs":fs_index(key),"fa":(first_acc[key].isoformat() if key in first_acc else ""),"dm":a.get("dm",""),
            "venc":a.get("venc",""),"ini":a.get("ini",""),"plano":a.get("plano",""),"termoMeses":a.get("termoMeses",0),"ult":(last_acc[key].isoformat() if key in last_acc else ""),"wk":week_acc.get(key,[0]*WEEKS_KEEP),
            "score":_sc["score"],"scoreBand":_sc["band"],"scorePrio":_sc["prio"],"scoreConf":_sc["conf"],"scoreReasons":_sc["reasons"],"scoreJustif":(_sc["justif"] or ""),
            "foto":a.get("foto",""),"prof":a.get("prof",""),"profRole":a.get("profRole",""),
            "email":a.get("email",""),"tel":a.get("tel",""),"novoCad":_is_novo,
        })
students.sort(key=lambda s:(s["u"],int(s["mat"]) if str(s["mat"]).isdigit() else 0))

# ==== CALIBRAÇÃO: PRIORIDADE POR CAPACIDADE (top-K por unidade) ====
# P0 = a fila que a equipe consegue ligar na semana (~5% da base ativa da unidade ≈ capacidade:
# 4 consultoras × 2h/dia × 5 dias ÷ 12 min ≈ 200/semana ≈ 40/unidade). Rankeia por SCORE DENTRO da
# unidade (não por corte absoluto) → volume estável mesmo com o score inflando no mês parcial.
# Piso de score evita "inventar" P0 em unidade calma. Tudo calibrável por env.
P0_FRAC  = float(os.environ.get("PACTO_P0_FRAC",  "0.05"))
P1_FRAC  = float(os.environ.get("PACTO_P1_FRAC",  "0.12"))
P2_FRAC  = float(os.environ.get("PACTO_P2_FRAC",  "0.25"))
P0_FLOOR = int(os.environ.get("PACTO_P0_FLOOR", "45"))
_by_unit = defaultdict(list)
for _s in students:
    _by_unit[_s["u"]].append(_s)
for _u, _lst in _by_unit.items():
    _act = [s for s in _lst if not s.get("scoreJustif")]   # acionáveis (contrato ativo, não justificado)
    _act.sort(key=lambda s: -s.get("score", 0))
    _n = len(_act)
    _k0 = int(round(_n*P0_FRAC)); _k1 = int(round(_n*P1_FRAC)); _k2 = int(round(_n*P2_FRAC))
    for _i, _s in enumerate(_act):
        if _i < _k0 and _s.get("score", 0) >= P0_FLOOR: _s["scorePrio"] = "P0"
        elif _i < _k0 + _k1:              _s["scorePrio"] = "P1"
        elif _i < _k0 + _k1 + _k2:        _s["scorePrio"] = "P2"
        else:                             _s["scorePrio"] = "P3"
    for _s in _lst:
        if _s.get("scoreJustif"): _s["scorePrio"] = "—"
print(f"[calib] prioridade por capacidade: P0={P0_FRAC:.0%}/unid (piso score {P0_FLOOR}), P1={P1_FRAC:.0%}, P2={P2_FRAC:.0%}", file=sys.stderr)

# ---- churn set-based (secao 8) ----
# attrs "achatado" por (unit,key): grupo/sexo/faixa nao mudam mes a mes. Resolve o caso em que a
# PERDA (por carencia) vem de um mes ANTERIOR ao 'pos' consultado -> antes o profile achava vazio
# (byCat/bySex/byBand saiam vazios em TODA transicao -> "Perfil de quem sai" vazio em todo filtro).
attrs_flat = {}
for _p in range(NMONTHS):
    for _uk, _a in attrs.get(_p, {}).items():
        attrs_flat[_uk] = _a   # ultima posicao vence (dados mais recentes do aluno)
def profile(mats, pos, unit):
    cat=Counter(); sex=Counter(); bnd=Counter()
    for mat in mats:
        a=attrs_flat.get((unit,mat))
        if not a: continue
        cat[a["grupo"]]+=1; sex[a["sexo"]]+=1; bnd[a["band"]]+=1
    return dict(cat),dict(sex),dict(bnd)
def merge(d,e):
    for k,v in e.items(): d[k]=d.get(k,0)+v
def pre(lst):
    n=len(lst); mean=round(sum(lst)/n,1) if n else 0
    zero=round(100*sum(1 for x in lst if x==0)/n) if n else 0
    return mean,zero
# ---- Tempo de permanencia (tenure): dm -> mes de referencia (em meses) ----
def _tenure_m(dm_iso, ref):
    d=parse_dt(dm_iso)
    if not d: return None
    return max(0,(ref.year-d.year)*12+(ref.month-d.month))
TENURE_BANDS=["0–3m","3–6m","6–12m","1–2 anos","2+ anos"]
def _tband(m):
    if m<3: return TENURE_BANDS[0]
    if m<6: return TENURE_BANDS[1]
    if m<12: return TENURE_BANDS[2]
    if m<24: return TENURE_BANDS[3]
    return TENURE_BANDS[4]
def tenure_profile(mats, unit, ref):
    c=Counter(); vals=[]
    for mat in mats:
        a=attrs_flat.get((unit,mat))
        if not a: continue
        tm=_tenure_m(a.get("dm",""), ref)
        if tm is None: continue
        vals.append(tm); c[_tband(tm)]+=1
    return dict(c), (round(sum(vals)/len(vals),1) if vals else 0)
def motivo_profile(mats, unit):
    # motivo de saida (situacaoContrato) de quem saiu. "" -> "Sem informação" (agrupado).
    c=Counter()
    for mat in mats:
        a=attrs_flat.get((unit,mat))
        mv=(a.get("motivo") if a else "") or "Sem informação"
        c[mv]+=1
    return dict(c)

trans_pairs=[(i,i+1) for i in range(NMONTHS-1)]
# CARENCIA: "perda" so conta apos N meses de inatividade real (1 mes de lapso e perdoado).
# Base com carencia = ativo no mes OU nos (N-1) meses anteriores. Mantem a identidade de
# conjuntos (base+novos-perdas=base seguinte), entao a checagem de consistencia continua fechando.
CARENCIA = int(os.environ.get("PACTO_CARENCIA", "2"))
def act_g(m, u):
    s = set()
    for d in range(max(1, CARENCIA)):
        s |= active.get((m - d, u), set())
    return s
churn={}; unit_data={}
for unit in UNIT_KEYS:
    trans=[]; ev=[]; ret=[]
    for a,b in trans_pairs:
        A=act_g(a,unit); B=act_g(b,unit)   # base com carencia (nao mais o snapshot cru)
        perdas=A-B; novos=B-A; retidos=A&B
        cC,cS,cB=profile(perdas,a,unit)
        _bref=datetime.date(ORDERED[b][0],ORDERED[b][1],1)   # mes em que saiu -> tenure = dm ate aqui
        cT,tMean=tenure_profile(perdas,unit,_bref)
        cMo=motivo_profile(perdas,unit)
        # prazo comprometido (meses) de quem entrou/saiu: soma dos termoMeses do plano de cada aluno.
        # Espelha o "comprometido em contrato" da aba de risco (default 12 p/ manter coerencia).
        _sumTermo=lambda ms: sum((attrs_flat.get((unit,m),{}).get("termoMeses") or 12) for m in ms)
        nTermo=_sumTermo(novos); pTermo=_sumTermo(perdas)
        trans.append({"de":MESES[a],"para":MESES[b],"perdas":len(perdas),"novos":len(novos),
                      "transf":0,"transfIn":0,"retidos":len(retidos),"base":len(A),
                      "byCat":cC,"bySex":cS,"byBand":cB,"byTenure":cT,"tenureMean":tMean,"byMotivo":cMo,
                      "novosTermo":nTermo,"perdasTermo":pTermo})
        # pre-perda usa ACESSO da catraca -> cego de catraca fica de fora (nao tem sinal de acesso)
        for mat in perdas:
            if (unit,mat) not in FREQ_BLIND: ev.append(acc[(unit,a)].get(mat,0))
        for mat in retidos:
            if (unit,mat) not in FREQ_BLIND: ret.append(acc[(unit,a)].get(mat,0))
    cm,cz=pre(ev); rm,rz=pre(ret)
    churn[unit]={"trans":trans,"pre":{"churnMean":cm,"churnZeroPct":cz,"retMean":rm,"retZeroPct":rz}}
    unit_data[unit]=(trans,ev,ret)
rede=[]; rede_ev=[]; rede_ret=[]
for i,(a,b) in enumerate(trans_pairs):
    perdas=novos=retidos=base=0; cC={};cS={};cB={};cT={};cMo={};_twsum=0.0; nTermoR=0; pTermoR=0
    for unit in UNIT_KEYS:
        t=unit_data[unit][0][i]
        perdas+=t["perdas"]; novos+=t["novos"]; retidos+=t["retidos"]; base+=t["base"]
        merge(cC,t["byCat"]); merge(cS,t["bySex"]); merge(cB,t["byBand"]); merge(cT,t.get("byTenure",{})); merge(cMo,t.get("byMotivo",{}))
        _twsum += (t.get("tenureMean",0) or 0)*t["perdas"]
        nTermoR += t.get("novosTermo",0); pTermoR += t.get("perdasTermo",0)
    rede.append({"de":MESES[a],"para":MESES[b],"perdas":perdas,"novos":novos,"transf":0,"transfIn":0,
                 "retidos":retidos,"base":base,"byCat":cC,"bySex":cS,"byBand":cB,
                 "byTenure":cT,"tenureMean":round(_twsum/perdas,1) if perdas else 0,"byMotivo":cMo,
                 "novosTermo":nTermoR,"perdasTermo":pTermoR})
for unit in UNIT_KEYS: rede_ev+=unit_data[unit][1]; rede_ret+=unit_data[unit][2]
cm,cz=pre(rede_ev); rm,rz=pre(rede_ret)
churn["REDE"]={"trans":rede,"pre":{"churnMean":cm,"churnZeroPct":cz,"retMean":rm,"retZeroPct":rz}}

# ---- consistencia base(a)+novos-perdas==base(b) ----
warns=[]
for unit in ["REDE"]+UNIT_KEYS:
    tr=churn[unit]["trans"]
    for i in range(len(tr)-1):
        exp=tr[i]["base"]+tr[i]["novos"]-tr[i]["perdas"]
        if exp!=tr[i+1]["base"]:
            warns.append(f"{unit} {tr[i]['de']}->{tr[i]['para']}: {exp}!={tr[i+1]['base']}")
print(("[WARN] consistencia churn:\n  "+"\n  ".join(warns)) if warns else "[ok] consistencia churn fecha", file=sys.stderr)

# ---- Fundacao v2: fluxos retido/saiu/novo/voltou (usa ledger/ym_of definidos acima) ----
def flows_for(scope):
    units = UNIT_KEYS if scope=="REDE" else [scope]
    rows=[]
    for a,b in trans_pairs:
        ymb=ym_of(b); A=set(); B=set()
        for u in units: A|=active[(a,u)]; B|=active[(b,u)]
        saiu=A-B; entrou=B-A; retido=A&B
        # Entrada em 2 vias (interino): NOVA MATRICULA (1a vez na rede, pelo ledger) | RECORRENTE
        # (ja teve cadastro antes -> renovacao OU retorno juntos). Separar renovacao de retorno
        # exige o HISTORICO de contratos (lapso real) — com o contrato vigente unico nao da.
        nova=sum(1 for k in entrou if ledger.get(_kh(k), ymb) >= ymb)
        recor=len(entrou)-nova
        rows.append({"de":MESES[a],"para":MESES[b],"base":len(A),
                     "retido":len(retido),"saiu":len(saiu),
                     "novo":nova,"recorrente":recor,"voltou":recor,   # voltou=recorrente (compat)
                     "churnPct":round(100*len(saiu)/len(A),1) if A else 0,
                     "retPct":round(100*len(retido)/len(A),1) if A else 0})
    return rows
flow={u:flows_for(u) for u in UNIT_KEYS}; flow["REDE"]=flows_for("REDE")

# ---- GATE DE VALIDACAO DE DADOS (premissa: nunca publicar dado quebrado) ----
def all_active(m):
    s=set()
    for unit in UNIT_KEYS: s |= active[(m,unit)]
    return s
MIN_ACTIVE=500; MIN_OVERLAP=0.60; MIN_ACC=500
# Os gates de integridade so ABORTAM para os meses RECENTES (janela historica de 2022+
# tem, de forma legitima, chain menor, mais churn e catraca parcial nos meses antigos).
# Meses fora da janela recente viram nota informativa ([hist]), nunca erro.
GATE_MONTHS = int(os.environ.get("PACTO_GATE_MONTHS", "6"))
GATE_START = max(0, NMONTHS - GATE_MONTHS)   # so enforce dos ultimos GATE_MONTHS meses
month_active=[all_active(m) for m in range(NMONTHS)]
errs=[]; hist=[]
# mes parcial (mes corrente, em curso ate D-1): seus avisos NAO abortam o publish
# (naturalmente tem menos acessos e overlap diferente por estar incompleto).
_today_g = datetime.date.today()
PARTIAL_M = base_pos if (tuple(ORDERED[base_pos]) == (_today_g.year, _today_g.month)) else -1
if PARTIAL_M >= 0:
    print(f"[info] gate: mes-base {MESES[PARTIAL_M]} e' parcial (em curso) -> avisos dele viram nota, nao abortam", file=sys.stderr)
def _flag(m, msg):
    (errs if (m >= GATE_START and m != PARTIAL_M) else hist).append(msg)
for m in range(NMONTHS):
    if len(month_active[m])<MIN_ACTIVE:
        _flag(m, f"mes {MESES[m]}: {len(month_active[m])} alunos ativos (<{MIN_ACTIVE})")
for m in range(NMONTHS-1):
    A,B=month_active[m],month_active[m+1]
    ratio=len(A&B)/(min(len(A),len(B)) or 1)
    if ratio<MIN_OVERLAP:
        _flag(PARTIAL_M if (m+1==PARTIAL_M) else m, f"overlap REDE {MESES[m]}->{MESES[m+1]}: {ratio:.0%} (<{MIN_OVERLAP:.0%})")
for unit in UNIT_KEYS:
    for m in range(NMONTHS-1):
        A=active[(m,unit)]; B=active[(m+1,unit)]
        if not A or not B: continue
        ratio=len(A&B)/(min(len(A),len(B)) or 1)
        if ratio<MIN_OVERLAP:
            _flag(PARTIAL_M if (m+1==PARTIAL_M) else m, f"overlap {unit} {MESES[m]}->{MESES[m+1]}: {ratio:.0%} (<{MIN_OVERLAP:.0%})")
for m in range(NMONTHS):
    tot=sum(sum(acc[(unit,m)].values()) for unit in UNIT_KEYS)
    if tot<MIN_ACC:
        _flag(m, f"acessos {MESES[m]}: {tot} (<{MIN_ACC})")
if hist:
    print(f"[hist] {len(hist)} avisos em meses antigos (fora dos ultimos {GATE_MONTHS}; nao abortam):\n  "+"\n  ".join(hist), file=sys.stderr)
if errs:
    print("[ERRO VALIDACAO] dados nao confiaveis nos meses recentes, NAO publicando:\n  "+"\n  ".join(errs), file=sys.stderr)
    sys.exit(1)
print(f"[ok] gate de validacao passou (recentes: ultimos {GATE_MONTHS} meses)", file=sys.stderr)

meta={}
try: meta=json.load(open(os.path.join(DATA_DIR,"meta.json")))
except Exception: pass

# ---- TICKET dinamico: faturamento real (API, ultimo mes fechado) / alunos ativos ----
_ativos_un = {}
for _s in students: _ativos_un[_s["u"]] = _ativos_un.get(_s["u"], 0) + 1
tickets_out = dict(TICKETS); ticket_mes = ""  # fallback = valores fixos
try:
    _fj = json.load(open(os.path.join(DATA_DIR, "faturamento.json")))
    _fat = _fj.get("faturamento", {})
    for _u in UNIT_KEYS:
        if _fat.get(_u) and _ativos_un.get(_u):
            tickets_out[_u] = round(float(_fat[_u]) / _ativos_un[_u], 2)
    _ym = _fj.get("mes", "")
    if _ym and "-" in _ym:
        _yy, _mm = _ym.split("-"); ticket_mes = f"{ABBR[int(_mm)]}/{_yy}"
    print(f"[ticket] dinamico (faturamento {_fj.get('mes')}): {tickets_out}", file=sys.stderr)
except Exception as _e:
    print(f"[ticket] valores fixos (sem faturamento.json): {_e}", file=sys.stderr)

# basePartial: o mes-base (mais recente) e' o mes corrente do calendario? -> ultimo mes e' parcial (D-1)
_today = datetime.date.today()
BASE_PARTIAL = (tuple(ORDERED[base_pos]) == (_today.year, _today.month))
print(f"[info] basePartial={BASE_PARTIAL} (mes-base {ORDERED[base_pos]} vs hoje {(_today.year,_today.month)})", file=sys.stderr)

# ==== Onda 3: BACKTEST MENSAL — o sinal mensal (parada/queda) antecipa o churn? ====
# Churn(m): ativo em m e AUSENTE da base em m+1 E m+2 (saída real, com carência).
# Flag(m): parou (acessos=0) OU caiu >=50% vs média dos 2 meses anteriores.
def _acc_of(unit, m, k): return acc.get((unit, m), {}).get(k, 0)
def _flag_at(unit, m, k):
    a = _acc_of(unit, m, k)
    if a == 0: return True
    b3 = [_acc_of(unit, mm, k) for mm in range(max(0, m-2), m)]
    med = (sum(b3)/len(b3)) if b3 else 0
    return med > 0 and a < 0.5*med
_TP=_FN=_FP=_TN=0; _leads=[]
for _unit in UNIT_KEYS:
    for _m in range(NMONTHS-2):
        _base = active.get((_m, _unit), set())
        _nx1 = active.get((_m+1, _unit), set()); _nx2 = active.get((_m+2, _unit), set())
        for _k in _base:
            if (_unit,_k) in FREQ_BLIND: continue   # cego de catraca: backtest mede o preditor por ACESSO
            _churn = (_k not in _nx1) and (_k not in _nx2)
            _fl = _flag_at(_unit, _m, _k)
            if   _churn and _fl:       _TP+=1
            elif _churn and not _fl:   _FN+=1
            elif (not _churn) and _fl: _FP+=1
            else:                      _TN+=1
            if _churn and _fl:
                _lead=1
                for _b in range(1,6):
                    if _m-_b>=0 and _flag_at(_unit,_m-_b,_k): _lead+=1
                    else: break
                _leads.append(_lead)
_prec = _TP/(_TP+_FP) if (_TP+_FP) else 0
_rec  = _TP/(_TP+_FN) if (_TP+_FN) else 0
_leadm = (sum(_leads)/len(_leads)) if _leads else 0
backtest = {"gerado": datetime.date.today().isoformat(), "metodo": "mensal (parada ou queda >=50%)",
            "TP":_TP, "FN":_FN, "FP":_FP, "TN":_TN,
            "precisao": round(_prec,3), "recall": round(_rec,3),
            "antecedenciaMesesMedia": round(_leadm,2),
            "churns": _TP+_FN, "alertas": _TP+_FP}
print(f"[backtest] mensal: precisao={_prec:.0%} recall={_rec:.0%} antecedencia={_leadm:.1f}m (churns={_TP+_FN} TP={_TP} FP={_FP})", file=sys.stderr)

# ==== Onda 4: CURVA DE SOBREVIVENCIA — retencao por coorte de ENTRADA (1a aparicao) ====
# De quem entrou no mes M, quantos % seguem ATIVOS em M, M+1, M+2...? Cresce sozinha com os meses.
_coh_mem = defaultdict(set)
for _k, _ym in first_seen_ym.items():
    _coh_mem[_ym].add(_k)
sobrevivencia = {}
for _ym, _mem in _coh_mem.items():
    _p = YM2POS.get(_ym)
    if _p is None: continue
    if _p == 0: continue    # JANELA: o 1o mes (Jan) e' CENSURADO A ESQUERDA (mistura veteranos que
                            # so aparecem agora) -> nao e' safra de entrada pura; fora da curva.
    _n0 = len(_mem)
    if _n0 < 10: continue   # ignora coortes minusculas (ruido)
    _curva = []
    for _d in range(0, NMONTHS - _p):
        _act = set()
        for _u in UNIT_KEYS: _act |= active.get((_p+_d, _u), set())
        _curva.append(round(len(_mem & _act)/_n0, 3))
    sobrevivencia[MESES[_p]] = {"n": _n0, "curva": _curva}
print(f"[sobrevivencia] {len(sobrevivencia)} coortes de entrada (retencao ao longo dos meses)", file=sys.stderr)

# ==== Onda 2b: EFETIVIDADE — cruza execucoes da Agenda (resultados.json) com a frequencia ====
# resultados.json = execucoes+iniciativas lidas do D1 (read-back no workflow, ANTES do build).
# Recuperou = aluno voltou a acessar DEPOIS do contato (ult >= data do contato). Sem PII no output.
efetividade = {"gerado": _TODAY.isoformat(), "temDados": False}
try:
    _res = json.load(open("resultados.json"))
    _rows = _res.get("results", _res) if isinstance(_res, dict) else _res
    if not isinstance(_rows, list): _rows = []
except Exception:
    _rows = []
if _rows:
    _byMat = {str(_s.get("mat","")): _s for _s in students}
    _tot=_cont=_rec=0; _receita=0.0
    _prio = defaultdict(lambda: [0,0])   # prioridade -> [alertas, recuperados]
    _cut = _TODAY - datetime.timedelta(days=42)   # efetividade das ultimas ~6 semanas (NAO acumulado)
    _seen = set()                                  # distinto por (matricula, semana) -> nao conta o mesmo alerta 2x
    for _r in _rows:
        _cd = parse_dt(_r.get("criado") or _r.get("registrado") or "")
        if _cd and _cd < _cut: continue            # fora do periodo recente
        _mat = str(_r.get("mat") or _r.get("matricula") or "")
        _dk = (_mat, str(_r.get("semana") or ""))
        if _dk in _seen: continue                  # alerta distinto por aluno+semana
        _seen.add(_dk)
        _tot += 1
        _contacted = (_r.get("status") or "") == "realizado"
        if _contacted: _cont += 1
        _stu = _byMat.get(_mat)
        _pr = _r.get("prio") or _r.get("prioridade") or "?"
        _prio[_pr][0] += 1
        # Recuperado = SO' entre os CONTATADOS, e so' se voltou a acessar DEPOIS do contato.
        # (antes contava recuperacao em todo alerta e dividia por contatados -> dava >100%, ex.: 8200%)
        _recovered = False
        if _contacted and _stu and _stu.get("ult"):
            _ud = parse_dt(_stu["ult"]); _ad = parse_dt(_r.get("registrado") or "")
            if _ud and _ad and _ud > _ad: _recovered = True
        if _recovered:
            _rec += 1; _prio[_pr][1] += 1
            if _stu: _receita += tickets_out.get(_stu.get("u"), 0)
    efetividade = {"gerado": _TODAY.isoformat(), "temDados": True, "janelaDias": 42, "poucosDados": (_cont < 5),
        "alertas": _tot, "contatados": _cont, "taxaContato": round(_cont/_tot,3) if _tot else 0,
        "recuperados": _rec, "taxaRecuperacao": round(_rec/_cont,3) if _cont else 0,
        "receitaPreservada": round(_receita,2),
        "porPrioridade": {k: {"alertas": v[0], "recuperados": v[1]} for k,v in _prio.items()}}
print(f"[efetividade] alertas={efetividade.get('alertas',0)} contatados={efetividade.get('contatados',0)} recuperados={efetividade.get('recuperados',0)} R$preservada={efetividade.get('receitaPreservada',0)}", file=sys.stderr)

out = {"students":students,"meses":MESES,"unidades":UNIDADES,"udps":UDPS,"backtest":backtest,"efetividade":efetividade,"sobrevivencia":sobrevivencia,
       "churn":churn,"tickets":tickets_out,"ticketNatal":TICKET_NATAL,"ticketMes":ticket_mes,
       "baseMonth":MESES[base_pos],"junMax":JUN_MAX,"flow":flow,"basePartial":BASE_PARTIAL,
       "weekRef":REF_MON.isoformat(),"weeksKeep":WEEKS_KEEP,
       "lagoUnit":LAGO_UNIT,"lagoExcluded":len(LAGO_EXCLUDED),
       "lagoFitCats":sorted(LAGO_FIT_CATS),
       "natalUnit":("Natal" if os.environ.get("PACTO_ENABLE_NATAL")=="1" else ""),
       "natalFreq":NATAL_FREQ,"natalChurnStart":"Mar.26","natalCatraca":"01/07/2026",
       "baseUpdated":meta.get("baseUpdated",""),
       "baseUpdatedBy":meta.get("baseUpdatedBy","") or meta.get("baseUpdatedByName","")}
print(f"[lago] excluidos por nao passar pela catraca (Agua/Luta puros): {len(LAGO_EXCLUDED)}", file=sys.stderr)
with open(os.path.join(DATA_DIR,"freq_multi.json"),"w") as f:
    json.dump(out,f,ensure_ascii=False,separators=(", ",": "))
print(f"[ok] freq_multi.json: {len(students)} students; baseMonth={MESES[base_pos]}; junMax={JUN_MAX}", file=sys.stderr)

# ---- PONTE App Treino: feed slim de PRESENCA por aluno, SEM PII (matricula HASHEADA) ----
# Publicado na raiz do repo (publico) -> o CRM App Treino consome via raw.githubusercontent
# e cruza pelo mesmo hash. So expomos: ultima visita (recencia presencial) + acessos/semana (tendencia).
import hashlib as _hl
def _pmh(m):   # hash estavel da matricula normalizada (16 hex) — identico dos dois lados
    return _hl.sha256(str(m).encode("utf-8")).hexdigest()[:16]
_presenca = {}
for _s in students:
    _m = _s.get("mat")
    if not _m:
        continue
    _presenca[_pmh(_m)] = {"ult": _s.get("ult",""), "wk": _s.get("wk",[])}
_feed = {"gerado": datetime.date.today().isoformat(), "baseMonth": MESES[base_pos],
         "weekRef": REF_MON.isoformat(), "weeksKeep": WEEKS_KEEP, "alunos": _presenca}
with open("presenca.json","w") as _f:   # raiz do repo (nao gitignorado)
    json.dump(_feed,_f,ensure_ascii=False,separators=(",",":"))
print(f"[ponte] presenca.json: {len(_presenca)} alunos (matricula hasheada, sem PII)", file=sys.stderr)

# ---- Onda 3: SCORE_HISTORY — snapshot SEMANAL de score/prioridade por aluno (raiz, commitado) ----
# Rotulo do backtest: "qual era o score quando o alerta saiu". Matricula HASHEADA (mesmo _pmh, sem PII).
# 1 snapshot por semana ISO (idempotente: reruns na mesma semana sobrescrevem). Poda p/ HIST_WEEKS semanas.
SCORE_HIST_PATH = "score_history.json"
HIST_WEEKS = int(os.environ.get("PACTO_SCORE_HIST_WEEKS", "26"))
try:
    _hist = json.load(open(SCORE_HIST_PATH))
    if not isinstance(_hist, dict): _hist = {}
except Exception:
    _hist = {}
_wk_key = REF_MON.isoformat()   # segunda-feira da semana corrente = chave do snapshot
_snap = {}
for _s in students:
    _m = _s.get("mat")
    if not _m: continue
    _snap[_pmh(_m)] = [_s.get("score", 0), _s.get("scorePrio", "")]
_hist[_wk_key] = _snap          # idempotente: sobrescreve a semana corrente
for _old in sorted(_hist.keys())[:-HIST_WEEKS]:   # poda: mantem as ultimas HIST_WEEKS semanas
    del _hist[_old]
with open(SCORE_HIST_PATH, "w") as _f:
    json.dump(_hist, _f, ensure_ascii=False, separators=(",", ":"))
print(f"[hist] score_history.json: semana {_wk_key}, {len(_snap)} alunos, {len(_hist)} semanas retidas", file=sys.stderr)
