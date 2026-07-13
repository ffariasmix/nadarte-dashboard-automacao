#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pacto_fetch.py — Coletor PACTO (ZillyonWeb) -> arquivos no formato do motor.

Substitui/complementa o drive_download.py: para cada unidade, puxa da API
  1) roster ATIVO (dedupe por codigoCliente),
  2) /clientes/{matricula}/dados-pessoais  -> codPessoa + CPF + dataMatricula + demografia,
  3) /acessos-cliente/by-pessoa/{codPessoa} -> historico de acessos (inclui facial),
e ESCREVE, em <data_dir>, planilhas .bin (xlsx) + manifest.json + meta.json IDENTICAS
ao que o build_freq_multi.py ja le hoje (Alunos Ativos por mes; Acessos Catraca por unidade).
A juncao aluno x acesso e por CPF (chave que existe nas duas bases).

Modos:
  - producao:   le PACTO_KEY_<UNIT> do ambiente (GitHub Secrets). 1 chave por unidade.
  - PACTO_SELFTEST=1: nao chama API; gera dados sinteticos no MESMO formato (teste de formato/gate).
  - PACTO_ONLY=716Norte: limita a coleta a 1 unidade (canario).

Uso: python3 pacto_fetch.py <data_dir>
PII fica apenas nos .bin (gitignored/efemeros no CI), nunca em log.
"""
import os, sys, io, json, time, random, datetime, calendar, re
import urllib.request, urllib.error
from concurrent.futures import ThreadPoolExecutor
import openpyxl

WORKERS = int(os.environ.get("PACTO_WORKERS", "2"))   # gentil por chave (evita rate-limit)
RETRY_ROUNDS = int(os.environ.get("PACTO_RETRY_ROUNDS", "4"))
DRIVE_MONTHS = set()   # meses (yr,mn) que o Drive JA cobre em catraca (API nao repete)
_SHEET_MONTH = re.compile(r"^\s*(\d{1,2})\.\s*\w+\.?(\d{4})?")

def _sheet_month(sn):
    m = _SHEET_MONTH.match(str(sn))
    if not m:
        return None
    yr = int(m.group(2)) if m.group(2) else datetime.date.today().year
    return (yr, int(m.group(1)))

def window_months(start, end):
    """lista de (yr,mn) de start ate end (inclusive)."""
    out = []
    y, m = start
    while (y, m) <= end:
        out.append((y, m))
        m += 1
        if m > 12:
            m = 1; y += 1
    return out

def active_in_month(ini, fim, yr, mn, situacao="", strict_ativo=False):
    """ATIVO no mes, alinhado ao criterio da PACTO (mesmo do painel Drive):
    - quem esta ATIVO hoje conta em todo mes desde que ja era cliente (nao importa se o
      contrato "venceu" — inclui vencidos/a vencer, como o flag ATIVO faz);
    - quem NAO esta mais ativo (saiu) e reconstruido pelo contrato (meses que o contrato cobriu),
      para o churn/retencao do passado ficar sem vies.
    """
    first = datetime.date(yr, mn, 1)
    last = datetime.date(yr, mn, calendar.monthrange(yr, mn)[1])
    # ainda nao era cliente naquele mes (entrou depois)
    if ini and ini > last:
        return False
    if str(situacao).upper() == "ATIVO":
        return True   # membro ativo hoje e ja era cliente em M
    if strict_ativo:
        return False  # mes-base: conta SO o flag ATIVO (nao expande por contrato)
    # inativo/saiu -> usa o contrato para saber ate quando esteve ativo
    if fim and fim < first:
        return False
    if not ini and not fim:
        return False
    return True

BASE = "https://apigw.pactosolucoes.com.br"
DATA_DIR = sys.argv[1] if len(sys.argv) > 1 else "data"
# catraca: emite meses a partir daqui. Alinhado ao recorte de "ativos" disponivel
# (o motor precisa de roster ATIVO em todo mes da timeline; senao o gate barra).
# Configuravel por env p/ crescer quando entrarem mais meses reais.
WINDOW_START = tuple(int(x) for x in os.environ.get("PACTO_WINDOW_START", "2026-01").split("-"))
NOW = datetime.date.today()
ABBR = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}

def _prev_month(ym):
    y, m = ym
    return (y - 1, 12) if m == 1 else (y, m - 1)

# fim da janela = MES CORRENTE (parcial, dados ate D-1). O mes corrente entra nos ARRAYS de
# acesso, mas NAO no filtro de 'win' (ver coleta_unidade): o conjunto de alunos e definido so
# por meses FECHADOS, ficando identico ao run sem parcial -> custo identico (os acessos do mes
# corrente ja sao baixados de qualquer forma). Foi isso que resolveu o timeout do #72.
# Override: PACTO_WINDOW_END=YYYY-MM (ex.: fechar so ate o mes completo anterior).
if os.environ.get("PACTO_WINDOW_END"):
    WINDOW_END = tuple(int(x) for x in os.environ["PACTO_WINDOW_END"].split("-"))
else:
    WINDOW_END = (NOW.year, NOW.month)
# o mes-base e' o mes corrente (parcial)? -> exclui esse mes do filtro de 'win'
BASE_IS_CURRENT = (WINDOW_END == (NOW.year, NOW.month))

# unidade -> (label do motor, nome do Secret). Natal fora ate ago/2026 (confirmar).
UNITS = [
    ("716Norte",  "716 Norte",  "PACTO_KEY_716NORTE"),
    ("905Sul",    "905 Sul",    "PACTO_KEY_905SUL"),
    ("604Norte",  "604 Norte",  "PACTO_KEY_604NORTE"),
    ("LagoNorte", "Lago Norte", "PACTO_KEY_LAGONORTE"),
    ("LagoSul",   "Lago Sul",   "PACTO_KEY_LAGOSUL"),
]

# ------------------------- HTTP -------------------------
def http_get(key, path, timeout=30, tries=5):
    for i in range(tries):
        req = urllib.request.Request(BASE + path, method="GET")
        req.add_header("Authorization", "Bearer " + key)
        req.add_header("Accept", "application/json")
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.status, r.read().decode("utf-8", "replace")
        except urllib.error.HTTPError as e:
            code = e.code
            body = e.read().decode("utf-8", "replace") if e.fp else ""
            if code in (429, 500, 502, 503, 504) and i < tries - 1:
                # backoff exponencial c/ jitter (rate-limit da PACTO)
                time.sleep(min(12.0, 0.8 * (2 ** i)) + random.uniform(0, 0.4)); continue
            return code, body
        except Exception as e:
            if i < tries - 1:
                time.sleep(min(8.0, 0.8 * (2 ** i))); continue
            return -1, str(e)
    return -1, ""

def gj(key, path):
    st, b = http_get(key, path)
    try:
        return st, json.loads(b)
    except Exception:
        return st, b

def lst(o):
    if isinstance(o, list):
        return o
    if isinstance(o, dict):
        for k in ("content","data","result","results","rows","list","items","clientes","acessos"):
            v = o.get(k)
            if isinstance(v, list):
                return v
            if isinstance(v, dict):
                for k2 in ("lista","content","items","rows"):
                    if isinstance(v.get(k2), list):
                        return v[k2]
    return []

def gv(d, *names):
    if not isinstance(d, dict):
        return None
    low = {k.lower(): k for k in d}
    for n in names:
        if n.lower() in low:
            return d[low[n.lower()]]
    return None

def unwrap(o):
    if isinstance(o, dict):
        c = o.get("content")
        if isinstance(c, dict):
            return c
        return o
    return {}

def to_date(v):
    if v is None:
        return None
    try:
        if isinstance(v, (int, float)) or (isinstance(v, str) and str(v).isdigit()):
            n = int(v)
            if n > 10_000_000_000:
                n //= 1000
            return datetime.datetime.fromtimestamp(n, tz=datetime.timezone.utc).date()
        s = str(v)[:19].replace("T", " ")
        for f in ("%Y-%m-%d %H:%M:%S","%Y-%m-%d","%d/%m/%Y %H:%M:%S","%d/%m/%Y"):
            try:
                return datetime.datetime.strptime(s, f).date()
            except Exception:
                pass
    except Exception:
        return None
    return None

# ------------------------- COLETA -------------------------
def roster_full(key):
    """TODOS os clientes (ativos+inativos+visitantes), dedup por codigoCliente.
    RAPIDO + ROBUSTO: le totalElements e pagina direcionado por esse total, re-tentando
    apenas as paginas que a PACTO devolve vazias (paginacao instavel). 2 passadas no maximo."""
    seen = set(); out = []
    total = None

    def add(o):
        n = 0
        for c in lst(o):
            if not isinstance(c, dict):
                continue
            cc = gv(c, "codigoCliente", "codigo")
            if cc is not None and cc not in seen:
                seen.add(cc); out.append(c); n += 1
        return n

    st, o = gj(key, "/clientes/simples?page=0&size=200")
    if isinstance(o, dict):
        total = gv(o, "totalElements", "total", "totalRegistros", "totalRows")
    add(o)
    npages = (int(total) // 200 + 3) if total else 300
    for sweep in range(2):
        for pg in range(1, npages):
            if total and len(out) >= int(total):
                break
            st, o = gj(key, f"/clientes/simples?page={pg}&size=200")
            r = lst(o)
            if not r:  # pagina instavel -> 1 re-tentativa
                st, o = gj(key, f"/clientes/simples?page={pg}&size=200")
            add(o)
        if total and len(out) >= int(total):
            break
    print(f"[roster] {len(out)}/{total} clientes coletados", file=sys.stderr)
    return out

FOTOS = os.environ.get("PACTO_FOTOS", "").strip().lower() in ("1", "true", "yes", "on")

def fotos_professores(key):
    """{prof_id: {'nome':..,'foto':..}} e {nome_upper: foto}. Guia 'Fotos'."""
    by_id, by_nome = {}, {}
    try:
        st, o = gj(key, "/psec/colaboradores/bi-professores-vinculos")
        if st == 200:
            for it in lst(o):
                p = gv(it, "professor") if isinstance(it, dict) else None
                if not isinstance(p, dict):
                    continue
                pid = gv(p, "id"); nome = gv(p, "nome") or ""; img = gv(p, "imageUri") or ""
                if pid is not None:
                    by_id[str(pid)] = {"nome": nome, "foto": img}
                if nome:
                    by_nome[str(nome).strip().upper()] = img
    except Exception:
        pass
    return by_id, by_nome

# Tipos de vinculo que contam como "professor" (modelo B). Consultor/comercial fica de fora.
PROF_TIPOS = ("PROFESSOR", "INSTRUTOR", "TREINADOR", "PERSONAL", "PROF")

def _prof_do_cliente(cli):
    """(nome, tipo) do vinculo de PROFESSOR/INSTRUTOR do aluno. Foto nao existe na API.
    Estrutura real: vinculos[].{codigo, colaborador{codigo,cpf,email,nome}, tipoVinculo}."""
    vincs = gv(cli, "vinculos")
    if not isinstance(vincs, list):
        return "", ""
    for v in vincs:
        if not isinstance(v, dict):
            continue
        tipo = str(gv(v, "tipoVinculo") or "").strip()
        if not any(k in tipo.upper() for k in PROF_TIPOS):
            continue                      # ignora CONSULTOR e outros vinculos nao-docentes
        colab = gv(v, "colaborador")
        nome = (gv(colab, "nome") or "") if isinstance(colab, dict) else ""
        if nome:
            return str(nome).strip(), (tipo.title() if tipo else "Professor")
    return "", ""

def fetch_client_full(key, c, wmonths, prof_by_id=None, prof_by_nome=None):
    """1 cliente -> (rec, [datas na janela]). rec: attrs + contrato (ini/fim/sit).
    dados-pessoais (codPessoa+cpf+demografia) + 1 chamada de acessos (+foto/prof se PACTO_FOTOS)."""
    prof_by_id = prof_by_id or {}; prof_by_nome = prof_by_nome or {}
    try:
        M = gv(c, "matricula"); cc = gv(c, "codigoCliente", "codigo"); nome = gv(c, "nome") or ""; sit = str(gv(c, "situacao") or "")
        ini = to_date(gv(c, "inicioContrato")); fim = to_date(gv(c, "fimContrato"))
        modc = gv(c, "categoria") or ""
        st, o = gj(key, f"/clientes/{M}/dados-pessoais")
        dp_ok = (st == 200)
        dp = unwrap(o) if dp_ok else {}
        cp = gv(dp, "codigoPessoa", "codPessoa"); cpf = gv(dp, "cpf") or ""
        # MODALIDADE real: vem da descricao do(s) contrato(s), nao do cadastro (que vem vazio).
        # Guia "Pacto API - Modalidades e Categorias": GET /v1/contrato/matricula/{mat} -> content[].descricao
        # Escopo necessario na chave: adm:cadastros:contrato:modelos-de-contrato:consultar
        mod_txt = ""; contract_ok = True
        st3, o3 = gj(key, f"/v1/contrato/matricula/{M}")
        if st3 == 200:
            descs = [str(gv(it, "descricao") or "").strip() for it in lst(o3)]
            mod_txt = " , ".join(d for d in descs if d)
        elif st3 in (429, 500, 502, 503, 504):
            contract_ok = False          # transitorio (rate-limit/erro) -> re-tentar
        # 401/403 (sem escopo) e 404 (sem contrato) -> aceita vazio, sem re-tentar em loop
        # FOTO do aluno + professor vinculado (guia 'Pacto API - Fotos'). So se PACTO_FOTOS=1.
        foto = ""; prof = ""; prof_role = ""
        if FOTOS and cc is not None:
            st4, o4 = gj(key, f"/v1/cliente/{cc}")
            if st4 == 200:
                cli = unwrap(o4) if isinstance(o4, dict) else {}
                pes = gv(cli, "pessoa") if isinstance(gv(cli, "pessoa"), dict) else {}
                foto = gv(pes, "fotoUrl") or gv(dp, "urlFoto") or ""
                prof, prof_role = _prof_do_cliente(cli)
        rec = {
            "ulabel": None, "mat": M, "nome": nome, "cpf": cpf,
            "nasc": to_date(gv(dp, "dataNascimento", "datanasc", "nascimento")),
            "sexo": gv(dp, "sexo") or "",
            "mod": mod_txt or gv(dp, "categoria") or modc,
            "dm": to_date(gv(dp, "dataMatricula")) or ini,
            "ini": ini, "fim": fim, "sit": sit,
            "foto": foto, "prof": prof, "profRole": prof_role,
        }
        dates = []
        acc_ok = True
        if cp:
            st2, o2 = gj(key, f"/acessos-cliente/by-pessoa/{cp}?page=0&size=1000")
            acc_ok = (st2 == 200)
            for a in (lst(o2) if acc_ok else []):
                # by-pessoa pode trazer acessos de IRMAOS (mesmo codPessoa) -> so os DESTE cliente
                cl = gv(a, "cliente"); clcod = gv(cl, "codigo") if isinstance(cl, dict) else None
                if clcod is not None and cc is not None and str(clcod) != str(cc):
                    continue
                d = to_date(gv(a, "dtHrEntrada") or gv(a, "dataDeAcesso") or gv(a, "dataRegistro"))
                if d and (d.year, d.month) in wmonths:
                    dates.append(d)
        ok = dp_ok and acc_ok and contract_ok   # qualquer chamada transitoria falhou -> re-tentar
        return rec, dates, ok
    except Exception:
        return {"ulabel": None, "mat": gv(c, "matricula"), "nome": gv(c, "nome") or "", "cpf": "",
                "nasc": None, "sexo": "", "mod": "", "dm": None,
                "ini": to_date(gv(c, "inicioContrato")), "fim": to_date(gv(c, "fimContrato")),
                "sit": str(gv(c, "situacao") or "")}, [], False

def coleta_unidade(unit_key, unit_label, key):
    """FULL-API com FILA DE RE-TENTATIVA: quem falha (rate-limit) volta pra fila e e
    recoletado em ate RETRY_ROUNDS rodadas. Nada e descartado silenciosamente."""
    all_months = window_months(WINDOW_START, WINDOW_END)
    wmonths = set(all_months)   # janela completa (INCL. mes corrente parcial): arrays de acesso
    # BASE = so quem esteve ativo nos ultimos INCLUDE_MONTHS meses (evita puxar todo mundo que
    # passou pela rede desde 2022). O historico completo continua nos arrays desses alunos.
    INCLUDE_MONTHS = int(os.environ.get("PACTO_INCLUDE_MONTHS", "12"))
    # IMPORTANTE (fix do timeout #72): o mes corrente parcial NAO define o conjunto de alunos.
    # Como todo membro ATIVO conta em qualquer mes desde que entrou, incluir o mes corrente no
    # filtro incharia 'win' para toda a base ativa. Aqui 'win' e definido so por meses FECHADOS
    # -> identico ao run que funcionava. Os acessos do mes corrente ja vem no size=1000 e entram
    # via wmonths. Alunos que so entraram no mes corrente ficam de fora ate o mes fechar.
    closed_months = all_months[:-1] if (BASE_IS_CURRENT and len(all_months) > 1) else all_months
    recent = set(closed_months[-INCLUDE_MONTHS:]) if len(closed_months) > INCLUDE_MONTHS else set(closed_months)
    prof_by_id, prof_by_nome = (fotos_professores(key) if FOTOS else ({}, {}))
    if FOTOS:
        print(f"[fotos] {unit_key}: {len(prof_by_id)} professores mapeados", file=sys.stderr)
    full = roster_full(key)
    win = [c for c in full
           if any(active_in_month(to_date(gv(c, "inicioContrato")), to_date(gv(c, "fimContrato")), y, m, gv(c, "situacao"))
                  for (y, m) in recent)]
    print(f"[base] {unit_key}: {len(win)} alunos (ativos nos ultimos {INCLUDE_MONTHS}m) de {len(full)} no historico", file=sys.stderr)
    t0 = time.time()
    recs = []
    pending = list(win)
    rounds = 0
    while pending and rounds < RETRY_ROUNDS:
        rounds += 1
        failed = []
        with ThreadPoolExecutor(max_workers=WORKERS) as ex:
            for (rec, dates, ok), c in zip(ex.map(lambda c: fetch_client_full(key, c, wmonths, prof_by_id, prof_by_nome), pending), pending):
                if ok:
                    rec["ulabel"] = unit_label
                    recs.append((rec, dates))
                else:
                    failed.append(c)
        print(f"[t] {unit_key} rodada {rounds}: ok_acum={len(recs)} falhas={len(failed)} (base={len(full)} janela={len(win)})", file=sys.stderr)
        pending = failed
        if pending:
            time.sleep(8 * rounds)   # respira antes de re-tentar (rate-limit)
    # ultima tentativa p/ os que restaram: inclui o aluno (roster) mesmo sem acesso, p/ nao sumir do mes
    for c in pending:
        rec, dates, ok = fetch_client_full(key, c, wmonths, prof_by_id, prof_by_nome)
        rec["ulabel"] = unit_label
        recs.append((rec, dates))
    if pending:
        print(f"[t] {unit_key}: RESTARAM {len(pending)} sem coleta completa apos {rounds} rodadas", file=sys.stderr)
    print(f"[t] {unit_key}: FIM base={len(full)} janela={len(win)} recs={len(recs)} em {time.time()-t0:.0f}s", file=sys.stderr)
    return recs

# ------------------------- ESCRITA (formato do motor) -------------------------
AL_HEADER = ["MATRICULA","NOME","DOCUMENTO","NASCIMENTO","SEXO","MODALIDADE","DATA MATRICULA","VENCIMENTO","FOTO","PROF NOME","PROF TIPO"]
CT_HEADER = ["MAT. CLIENTE","NOME","CPF","DATA ENTRADA"]

def write_alunos_wb(path, unit_label_to_rows):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for label, rows in unit_label_to_rows.items():
        ws = wb.create_sheet(title=label[:31])
        ws.append(AL_HEADER)
        for r in rows:
            ws.append([r["mat"], r["nome"], r["cpf"], r["nasc"], r["sexo"], r["mod"], r["dm"], r.get("fim",""),
                       r.get("foto",""), r.get("prof",""), r.get("profRole","")])
    wb.save(path)

def write_catraca_wb(path, sheets):
    """sheets: {(year,month): [ {cpf,nome,data} ]}"""
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for (yr, mn) in sorted(sheets.keys()):
        ws = wb.create_sheet(title=f"{mn}. {ABBR[mn]}.{yr}"[:31])
        ws.append(CT_HEADER)
        for r in sheets[(yr, mn)]:
            ws.append([r.get("mat", ""), r["nome"], r["cpf"], r["data"]])
    wb.save(path)

def merge_catraca_into(path, sheets):
    """Acrescenta as folhas de mes da API dentro do workbook de catraca do Drive (mesmo arquivo),
    somente para meses ainda ausentes. Preserva o historico completo do Drive."""
    wb = openpyxl.load_workbook(io.BytesIO(open(path, "rb").read()))
    have = set()
    for sn in wb.sheetnames:
        mm = _sheet_month(sn)
        if mm:
            have.add(mm)
    add = 0
    for (yr, mn) in sorted(sheets.keys()):
        if (yr, mn) in have:
            continue
        ws = wb.create_sheet(title=f"{mn}. {ABBR[mn]}.{yr}"[:31])
        ws.append(CT_HEADER)
        for r in sheets[(yr, mn)]:
            ws.append([r.get("mat", ""), r["nome"], r["cpf"], r["data"]])
        add += 1
    wb.save(path)
    return add

# ------------------------- SELFTEST (sintetico) -------------------------
def selftest_data():
    """gera 5 unidades x 2 meses, ~640 ativos/unid, overlap ~85%, acessos suficientes."""
    random.seed(7)
    months = [(2026, 5), (2026, 6)]
    alunos_by_month = {m: {} for m in months}
    catraca_by_unit = {}
    for uk, ulabel, _ in UNITS:
        base_ids = list(range(1000, 1000 + 640))   # cpfs sinteticos
        prev = set()
        cat = {}
        for mi, m in enumerate(months):
            # 85% de retencao + alguns novos
            if mi == 0:
                cur = set(base_ids[:600])
            else:
                keep = set(random.sample(sorted(prev), int(len(prev) * 0.85)))
                novos = set(base_ids[600:640])
                cur = keep | novos
            prev = cur
            rows = []
            for cid in sorted(cur):
                cpf = f"{uk[:3].upper()}{cid:08d}"[:11].ljust(11, "0")
                rows.append({
                    "mat": str(cid), "nome": f"ALUNO {uk} {cid}", "cpf": cpf,
                    "nasc": datetime.date(1990, ((cid % 12) + 1), ((cid % 27) + 1)),
                    "sexo": "M" if cid % 2 else "F",
                    "mod": random.choice(["TRANSITO LIVRE","MUSCULACAO","NATACAO","MUAY THAI"]),
                    "dm": datetime.date(2024, 1, 1),
                })
                # acessos no mes (>=1 p/ maioria)
                naccess = random.choice([0, 2, 4, 8, 12])
                for _ in range(naccess):
                    day = random.randint(1, calendar.monthrange(m[0], m[1])[1])
                    cat.setdefault(m, []).append({"mat": str(cid), "cpf": cpf, "nome": f"ALUNO {uk} {cid}",
                                                  "data": datetime.date(m[0], m[1], day)})
            alunos_by_month[m][ulabel] = rows
        catraca_by_unit[(uk, ulabel)] = cat
    return months, alunos_by_month, catraca_by_unit

# ------------------------- MAIN -------------------------
def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    manifest = {}
    fid = 0
    def nextfid():
        nonlocal fid; fid += 1; return f"api{fid:04d}"

    if os.environ.get("PACTO_SELFTEST") == "1":
        months, alunos_by_month, catraca_by_unit = selftest_data()
        for (yr, mn) in months:
            f = nextfid(); path = os.path.join(DATA_DIR, f + ".bin")
            write_alunos_wb(path, alunos_by_month[(yr, mn)])
            manifest[f] = f"{mn}. Alunos Ativos Rede ({ABBR[mn]}.{yr})"
        for (uk, ulabel), sheets in catraca_by_unit.items():
            f = nextfid(); path = os.path.join(DATA_DIR, f + ".bin")
            write_catraca_wb(path, sheets)
            manifest[f] = f"Acessos Catrata Unidade {ulabel} (Nad'Arte)"
        json.dump(manifest, open(os.path.join(DATA_DIR, "manifest.json"), "w"), ensure_ascii=False)
        json.dump({"baseUpdated": NOW.isoformat(), "baseUpdatedBy": "selftest"},
                  open(os.path.join(DATA_DIR, "meta.json"), "w"), ensure_ascii=False)
        print(f"[selftest] escrito {len(manifest)} arquivos em {DATA_DIR}", file=sys.stderr)
        return

    # =================== SITPROBE: so cadastro (rapido) p/ achar o criterio de "ativo" ===================
    if os.environ.get("PACTO_SITPROBE") == "1":
        from collections import Counter
        ye, me = WINDOW_END
        first = datetime.date(ye, me, 1); last = datetime.date(ye, me, calendar.monthrange(ye, me)[1])
        def cov(c):  # contrato cobre o mes de referencia
            i = to_date(gv(c, "inicioContrato")); f = to_date(gv(c, "fimContrato"))
            if i and i > last: return False
            if f and f < first: return False
            return bool(i or f)
        def analyze(t):
            uk, ulabel, secret = t
            key = os.environ.get(secret, "").strip()
            if not key:
                print(f"[sit skip] {uk}: sem secret", file=sys.stderr); return None
            full = roster_full(key)
            dist = Counter(str(gv(c, "situacao") or "?").upper() for c in full)
            atv   = sum(1 for c in full if str(gv(c, "situacao") or "").upper() == "ATIVO")
            atr   = sum(1 for c in full if str(gv(c, "situacao") or "").upper() in ("ATIVO", "TRANCADO"))
            contr = sum(1 for c in full if cov(c))
            aoc   = sum(1 for c in full if str(gv(c, "situacao") or "").upper() == "ATIVO" or cov(c))
            print(f"[sit] {uk}: total={len(full)} dist={dict(dist)} | ATIVO={atv} ATIVO+TRANC={atr} contrato={contr} ATIVO|contrato={aoc}", file=sys.stderr)
            return (uk, atv, atr, contr, aoc)
        res = []
        with ThreadPoolExecutor(max_workers=5) as ex:
            for r in ex.map(analyze, UNITS):
                if r: res.append(r)
        labels = ["ATIVO", "ATIVO+TRANC", "contrato(cobre Jun)", "ATIVO|contrato"]
        for i, name in enumerate(labels):
            print(f"[REDE {name}] = {sum(x[i+1] for x in res)}  (alvo Drive Jun/26 = 3956)", file=sys.stderr)
        return

    # =================== FULL-API: reconstrucao por CONTRATO ===================
    # "Ativos por mes" reconstruidos pelo CONTRATO (inclui quem ja saiu) -> churn SEM vies.
    # Frequencia (facial) de TODOS que estiveram ativos na janela. Drive vira so fallback.
    wmonths = window_months(WINDOW_START, WINDOW_END)
    only = os.environ.get("PACTO_ONLY", "").strip()
    targets = [(uk, ulabel, secret) for (uk, ulabel, secret) in UNITS if not only or uk == only]

    def run_unit(t):
        uk, ulabel, secret = t
        key = os.environ.get(secret, "").strip()
        if not key:
            print(f"[skip] {uk}: sem secret {secret}", file=sys.stderr); return None
        try:
            return (uk, ulabel, coleta_unidade(uk, ulabel, key))
        except Exception as e:
            print(f"[ERRO] {uk}: {e}", file=sys.stderr); return None

    # unidades EM PARALELO (chaves independentes -> rate-limit por chave respeitado)
    results = []
    with ThreadPoolExecutor(max_workers=max(1, len(targets))) as ex:
        for r in ex.map(run_unit, targets):
            if r:
                results.append(r)

    alunos_by_month = {ym: {} for ym in wmonths}   # ym -> {ulabel: [rec]}
    catraca_by_unit = {}                            # ulabel -> {ym: [row]}
    for uk, ulabel, recs in results:
        cat = {}
        for ym in wmonths:
            # mes-base (WINDOW_END = mais recente) conta SO flag ATIVO; meses passados usam contrato (churn)
            _strict = (ym == WINDOW_END)
            alunos_by_month[ym][ulabel] = [
                rec for rec, _ in recs
                if active_in_month(rec["ini"], rec["fim"], ym[0], ym[1], rec["sit"], strict_ativo=_strict)
            ]
        for rec, dates in recs:
            for d in dates:
                cat.setdefault((d.year, d.month), []).append({"mat": rec["mat"], "cpf": rec["cpf"], "nome": rec["nome"], "data": d})
        catraca_by_unit[ulabel] = cat
        nwin = len(recs); ncpf = sum(1 for rec, _ in recs if len(str(rec["cpf"] or "")) >= 11)
        tot_acc = sum(len(v) for v in cat.values())
        por_mes = {f"{ym[0]}-{ym[1]:02d}": len(alunos_by_month[ym][ulabel]) for ym in wmonths}
        print(f"[ok] {uk}: janela={nwin} cpf={100*ncpf//max(1,nwin)}% acessos={tot_acc} ativos/mes={por_mes}", file=sys.stderr)

    # escrever: 1 arquivo de alunos por mes + 1 catraca por unidade
    for ym in wmonths:
        mes_rows = {ul: rows for ul, rows in alunos_by_month[ym].items() if rows}
        if mes_rows:
            f = nextfid(); write_alunos_wb(os.path.join(DATA_DIR, f + ".bin"), mes_rows)
            manifest[f] = f"{ym[1]}. Alunos Ativos Rede ({ABBR[ym[1]]}.{ym[0]})"
    for ulabel, cat in catraca_by_unit.items():
        if cat:
            f = nextfid(); write_catraca_wb(os.path.join(DATA_DIR, f + ".bin"), cat)
            manifest[f] = f"Acessos Catrata Unidade {ulabel} (Nad'Arte)"
    # so sobrescreve manifest se a API produziu dados; senao mantem o do Drive (fallback)
    if manifest:
        json.dump(manifest, open(os.path.join(DATA_DIR, "manifest.json"), "w"), ensure_ascii=False)
        json.dump({"baseUpdated": NOW.isoformat(), "baseUpdatedBy": "API PACTO"},
                  open(os.path.join(DATA_DIR, "meta.json"), "w"), ensure_ascii=False)
    print(f"[fim] full-API: {len(manifest)} arquivos, janela {wmonths[0]}..{wmonths[-1]}", file=sys.stderr)


if __name__ == "__main__":
    main()
