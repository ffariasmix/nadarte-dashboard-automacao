#!/usr/bin/env python3
# ============================================================
# build_ocupacao.py — Bloco Operacional (Ocupação) do Motor da Agenda.
#
# Agrega ENTRADAS de catraca por (unidade, tipo_de_dia, hora) e publica
# data/agenda_ocupacao.json com os slots OCIOSOS / de PICO por unidade.
#
# Fonte: os MESMOS .bin de catraca que o build_freq_multi.py lê
# (data/manifest.json). A coluna "Data entrada" é datetime completo
# (ex.: 2026-01-02 06:32:51) — usamos hora + dia-da-semana.
#
# CAVEAT HONESTO: a catraca só registra ENTRADA (sem saída). Portanto isto
# é proxy de PROCURA / FLUXO por horário, NÃO ocupação simultânea real.
# Serve muito bem para separar horários CHEIOS vs OCIOSOS.
#
# Limiar: relativo por unidade (quartis sobre o perfil seg–sex).
#         bottom 25% = ocioso · top 25% = pico.
# Volume: top-3 ociosos + top-2 picos por unidade (parametrizável).
# Janela: horário de funcionamento por unidade e tipo de dia (Nad'Arte).
#
# Uso (produção):  python3 build_ocupacao.py data
# Uso (teste local): OCUP_LOCAL_DIR="/caminho/Acessos Catraca" python3 build_ocupacao.py
# ============================================================
import sys, os, re, json, glob, io, datetime
from collections import defaultdict
import openpyxl

# Abre .xlsx OU .bin (xlsx com extensão trocada, como o pipeline usa): via BytesIO,
# que ignora a checagem de extensão do openpyxl. Mesmo truque do build_freq_multi.
def _load(path):
    return openpyxl.load_workbook(io.BytesIO(open(path, "rb").read()), read_only=True, data_only=True)

DATA_DIR   = sys.argv[1] if len(sys.argv) > 1 else "data"
WEEKS      = int(os.environ.get("OCUP_WEEKS", "8"))     # janela de semanas recentes
TOP_OCIOSO = int(os.environ.get("OCUP_TOP_OCIOSO", "3"))
TOP_PICO   = int(os.environ.get("OCUP_TOP_PICO", "2"))
# Fim de semana tem menos horas e menor volume → menos cards (1 ocioso + 1 pico por dia).
TOP_FDS_OCIOSO = int(os.environ.get("OCUP_TOP_FDS_OCIOSO", "1"))
TOP_FDS_PICO   = int(os.environ.get("OCUP_TOP_FDS_PICO", "1"))
LOCAL_DIR  = os.environ.get("OCUP_LOCAL_DIR")           # p/ teste sem manifest

UNIDADE_SLUG = {'716Norte':'716-norte','905Sul':'905-sul','604Norte':'604-norte',
                'LagoNorte':'lago-norte','LagoSul':'lago-sul'}

# Horário de funcionamento (h_ini inclusivo, h_fim exclusivo p/ a última hora cheia).
# tipo de dia: 'sem' = seg–sex · 'sab' = sábado · 'dom' = domingo.
HORARIOS = {
  '905-sul':    {'sem':(6,22), 'sab':(8,15), 'dom':(9,13)},
  '716-norte':  {'sem':(6,22), 'sab':(8,15), 'dom':(9,13)},
  '604-norte':  {'sem':(6,21), 'sab':(8,12)},              # domingo fechado
  'lago-norte': {'sem':(7,21), 'sab':(8,12)},              # domingo fechado
  'lago-sul':   {'sem':(6,21), 'sab':(8,13)},              # domingo fechado
}

def daytype(d):
    wd = d.weekday()
    return 'sem' if wd < 5 else ('sab' if wd == 5 else 'dom')

def parse_hdr(rows):
    """acha a linha de cabeçalho e o índice da coluna 'Data entrada'."""
    for i, r in enumerate(rows[:15]):
        if not r: continue
        for j, c in enumerate(r):
            if c and str(c).strip().lower().startswith("data entrada"):
                return i, j
    return None, None

# ---- descobrir arquivos de catraca ----
catraca = {}   # unit_key -> path
if LOCAL_DIR:
    for p in glob.glob(os.path.join(LOCAL_DIR, "**", "*.xlsx"), recursive=True):
        base = os.path.basename(p)
        if base.startswith("~$"): continue
        mc = re.search(r"Acessos Catr?ata Unidade (.+?) \(Nad", base)
        if mc:
            catraca[mc.group(1).strip().replace(" ", "")] = p
else:
    manifest = json.load(open(os.path.join(DATA_DIR, "manifest.json")))
    for fid, title in manifest.items():
        mc = re.search(r"Acessos Catr?ata Unidade (.+?) \(Nad", title)
        if mc:
            catraca[mc.group(1).strip().replace(" ", "")] = os.path.join(DATA_DIR, fid + ".bin")

assert catraca, "nenhum arquivo de catraca encontrado"
print(f"[ocup] catraca: {sorted(catraca.keys())}", file=sys.stderr)

# ---- 1ª passada: coletar todos os datetimes por unidade (achar âncora) ----
raw = defaultdict(list)   # unit_key -> [datetime,...]
for uk, path in catraca.items():
    if uk not in UNIDADE_SLUG:  # Natal etc. não têm catraca/slug
        continue
    wb = _load(path)
    for sn in wb.sheetnames:
        rows = list(wb[sn].iter_rows(values_only=True))
        hi, ci = parse_hdr(rows)
        if hi is None: continue
        for r in rows[hi+1:]:
            if not r or ci >= len(r): continue
            v = r[ci]
            if isinstance(v, datetime.datetime):
                raw[uk].append(v)
    wb.close()

# âncora = data mais recente observada (dados podem estar atrás de hoje)
todos = [v for lst in raw.values() for v in lst]
if not todos:
    print("[ocup] sem entradas datadas — abortando.", file=sys.stderr); sys.exit(0)
ancora = max(v.date() for v in todos)
corte  = ancora - datetime.timedelta(weeks=WEEKS)
print(f"[ocup] âncora={ancora} · janela={WEEKS}sem (corte {corte}) · entradas totais={len(todos)}", file=sys.stderr)

# ---- 2ª passada: contar entradas por (slug, daytype, hora) dentro da janela e do horário ----
cont = defaultdict(int)                 # (slug, dt, hora) -> nº entradas
dias_obs = defaultdict(set)             # (slug, dt) -> {datas distintas}  (p/ média por dia)
for uk, lst in raw.items():
    slug = UNIDADE_SLUG[uk]
    horas_un = HORARIOS.get(slug, {})
    for v in lst:
        d = v.date()
        if d < corte or d > ancora: continue
        dt = daytype(d)
        janela = horas_un.get(dt)
        if not janela: continue         # dia sem funcionamento (ex.: domingo fechado)
        h = v.hour
        if not (janela[0] <= h < janela[1]): continue
        cont[(slug, dt, h)] += 1
        dias_obs[(slug, dt)].add(d)

# ---- perfil por slot: média de entradas por dia + status (quartis por unidade/daytype) ----
def quartis(vals):
    s = sorted(vals); n = len(s)
    if n == 0: return (0, 0)
    q1 = s[max(0, int(round(0.25*(n-1))))]
    q3 = s[min(n-1, int(round(0.75*(n-1))))]
    return q1, q3

unidades_out = {}   # slug -> {dt: [{hora,media,status}]}
cards = []          # lista acionável (o que o motor consome)

for slug, horas_un in HORARIOS.items():
    unidades_out[slug] = {}
    for dt, (h0, h1) in horas_un.items():
        ndias_real = len(dias_obs.get((slug, dt), []))
        ndias = max(1, ndias_real)
        perfil = []
        for h in range(h0, h1):
            media = round(cont.get((slug, dt, h), 0) / ndias, 1)
            perfil.append({'hora': h, 'media': media})
        medias = [p['media'] for p in perfil]
        # dado suficiente? unidade/dia sem entradas (ou 1 dia só) não gera classificação nem card.
        temDado = (ndias_real >= 2) and any(m > 0 for m in medias)
        q1, q3 = quartis([m for m in medias if m > 0]) if temDado else (0, 0)
        for p in perfil:
            if temDado:
                if p['media'] <= q1:   p['status'] = 'ocioso'
                elif p['media'] >= q3: p['status'] = 'pico'
                else:                  p['status'] = 'normal'
            else:
                p['status'] = 'normal'
        unidades_out[slug][dt] = perfil

        # cards por dia (seg–sex + sáb/dom). Fim de semana usa top menor. Só com dado suficiente.
        if temDado:
            n_oc, n_pk = (TOP_OCIOSO, TOP_PICO) if dt == 'sem' else (TOP_FDS_OCIOSO, TOP_FDS_PICO)
            ociosos = sorted([p for p in perfil if p['status']=='ocioso' and p['media']>0], key=lambda x: x['media'])[:n_oc]
            picos   = sorted([p for p in perfil if p['status']=='pico'],   key=lambda x: -x['media'])[:n_pk]
            med_un  = round(sum(medias)/len(medias), 1) if medias else 0
            for p in ociosos:
                cards.append({'unidade':slug,'dia':dt,'hora':p['hora'],'media':p['media'],
                              'media_unidade':med_un,'status':'ocioso'})
            for p in picos:
                cards.append({'unidade':slug,'dia':dt,'hora':p['hora'],'media':p['media'],
                              'media_unidade':med_un,'status':'pico'})

out = {
  'gerado': datetime.date.today().isoformat(),
  'origem': 'ocupacao',
  'ancora': ancora.isoformat(),
  'janela_semanas': WEEKS,
  'aviso': 'entradas de catraca (sem saída) = proxy de procura/fluxo, não ocupação simultânea',
  'ocup': cards,             # acionável (motor)
  'unidades': unidades_out,  # perfil completo (transparência/debug)
}

destino = os.path.join(DATA_DIR, "agenda_ocupacao.json") if not LOCAL_DIR else "agenda_ocupacao.json"
with open(destino, "w") as f:
    json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
print(f"[ok] agenda_ocupacao.json: {len(cards)} cards (seg–sex {TOP_OCIOSO}+{TOP_PICO}; fim de semana {TOP_FDS_OCIOSO}+{TOP_FDS_PICO}/dia por unidade).", file=sys.stderr)
